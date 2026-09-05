"""Parser for ISA resident-foreigner statistics table 1."""

import calendar
import re
from pathlib import Path
from typing import Dict, Iterator, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from .errors import SchemaError
from .models import NationalityPopulationTotalRecord, PopulationRecord


REQUIRED_HEADERS = (
    "国籍・地域",
    "在留資格",
    "性別",
    "年齢（５歳階級）",
    "年齢",
    "都道府県",
    "在留外国人数",
)


def _split_label(value: object, separators: Sequence[str]) -> Tuple[str, str]:
    text = "" if value is None else str(value).strip()
    for separator in separators:
        if separator in text:
            code, label = text.split(separator, 1)
            return code.strip(), label.strip()
    return "", text


def _period_end_from_sheet_name(sheet_name: str) -> str:
    match = re.search(r"令和\s*(\d+)年\s*(\d+)月末", sheet_name)
    if not match:
        raise SchemaError("Could not derive period_end from sheet name: %s" % sheet_name)
    year = 2018 + int(match.group(1))
    month = int(match.group(2))
    day = calendar.monthrange(year, month)[1]
    return "%04d-%02d-%02d" % (year, month, day)


def _find_data_sheet(workbook) -> Tuple[object, int, Dict[str, int]]:
    required = set(REQUIRED_HEADERS)
    for worksheet in workbook.worksheets:
        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=30, values_only=True),
            start=1,
        ):
            values = [str(value).strip() if value is not None else "" for value in row]
            if required.issubset(set(values)):
                return worksheet, row_index, {
                    header: values.index(header) for header in REQUIRED_HEADERS
                }
    raise SchemaError("Population table 1 required columns were not found")


def _population_value(value: object) -> Tuple[Optional[int], bool]:
    if isinstance(value, str) and value.strip() == "-":
        return None, True
    if value in (None, ""):
        return None, False
    numeric = float(value)
    if not numeric.is_integer():
        raise SchemaError("Population count is not an integer: %r" % value)
    return int(numeric), False


def parse_population_t1(
    path: Path,
    *,
    source_id: str,
    period_end: Optional[str] = None,
) -> Iterator[PopulationRecord]:
    """Yield normalized rows without loading the complete official workbook into memory."""

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet, header_row, columns = _find_data_sheet(workbook)
        effective_period_end = period_end or _period_end_from_sheet_name(worksheet.title)
        for source_row, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if not any(value not in (None, "") for value in row):
                continue

            nationality_code, nationality = _split_label(
                row[columns["国籍・地域"]], ("：", ":")
            )
            status_code, status = _split_label(
                row[columns["在留資格"]], ("：", ":")
            )
            sex_code, sex = _split_label(row[columns["性別"]], ("：", ":"))
            age_group_code, age_group = _split_label(
                row[columns["年齢（５歳階級）"]], ("_",)
            )
            prefecture_code, prefecture = _split_label(
                row[columns["都道府県"]], ("：", ":")
            )
            value, suppressed = _population_value(row[columns["在留外国人数"]])

            required_labels = (nationality, status, sex, age_group, prefecture)
            if not all(required_labels):
                raise SchemaError("Incomplete population labels at source row %d" % source_row)

            yield PopulationRecord(
                period_end=effective_period_end,
                nationality_code=nationality_code,
                nationality=nationality,
                residence_status_code=status_code,
                residence_status=status,
                sex_code=sex_code,
                sex=sex,
                age_group_code=age_group_code,
                age_group=age_group,
                age=str(row[columns["年齢"]]).strip(),
                prefecture_code=prefecture_code,
                prefecture=prefecture,
                value=value,
                suppressed=suppressed,
                source_id=source_id,
                source_sheet=worksheet.title,
                source_row=source_row,
            )
    finally:
        workbook.close()


TOTAL_REGION_ALIASES = {
    "北米": "北アメリカ",
    "南米": "南アメリカ",
}
TOTAL_REGION_LABELS = {
    "アジア",
    "ヨーロッパ",
    "アフリカ",
    "北米",
    "北アメリカ",
    "南米",
    "南アメリカ",
    "オセアニア",
}


def _compact_label(value: object) -> str:
    return "".join(str(value or "").split())


def _canonical_total_region(label: str) -> str:
    return TOTAL_REGION_ALIASES.get(label, label)


def _total_period_end(sheet_name: str) -> str:
    match = re.match(r"^(\d{2})-(\d{2})-01", sheet_name.strip())
    if not match:
        raise SchemaError(
            "Could not derive nationality population period from sheet name: %s"
            % sheet_name
        )
    year = 2000 + int(match.group(1))
    month = int(match.group(2))
    day = calendar.monthrange(year, month)[1]
    return "%04d-%02d-%02d" % (year, month, day)


def _total_integer(value: object, *, source_row: int) -> int:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        raise SchemaError(
            "Missing nationality population at source row %d" % source_row
        )
    if not numeric.is_integer() or numeric < 0:
        raise SchemaError(
            "Invalid nationality population at source row %d" % source_row
        )
    return int(numeric)


def _validate_population_total_title(worksheet) -> None:
    title = _compact_label(
        " ".join(
            str(value)
            for row in worksheet.iter_rows(
                min_row=1, max_row=min(6, worksheet.max_row), values_only=True
            )
            for value in row
            if value not in (None, "")
        )
    )
    if not (
        ("第1表" in title or "第１表" in title)
        and "国籍・地域別" in title
        and "在留外国人" in title
    ):
        raise SchemaError("Workbook is not ISA resident-foreigner Table 1")


def _flat_total_header(worksheet) -> Optional[Tuple[int, Dict[str, int]]]:
    required = {"州", "国籍・地域", "在留資格", "在留外国人数"}
    for source_row, row in enumerate(
        worksheet.iter_rows(
            min_row=1, max_row=min(10, worksheet.max_row), values_only=True
        ),
        start=1,
    ):
        labels = [_compact_label(value) for value in row]
        if required.issubset(labels):
            return source_row, {label: labels.index(label) for label in labels if label}
    return None


def _make_population_total_record(
    *,
    period_end: str,
    source_region: Optional[str],
    source_nationality: str,
    population: int,
    source_id: str,
    source_sheet: str,
    source_row: int,
    source_column: int,
) -> NationalityPopulationTotalRecord:
    if source_region == "総数" and source_nationality == "総数":
        region = None
        nationality = None
        row_kind = "national_total"
    elif source_region == "無国籍" and source_nationality == "総数":
        region = None
        nationality = "無国籍"
        row_kind = "country_or_area"
    elif source_nationality.startswith("うち"):
        region = (
            _canonical_total_region(source_region)
            if source_region and source_region != "総数"
            else None
        )
        nationality = source_nationality
        row_kind = "subcategory"
    elif source_region in TOTAL_REGION_LABELS and source_nationality in (
        "総数",
        source_region,
    ):
        region = _canonical_total_region(source_region)
        nationality = None
        row_kind = "region_total"
    else:
        region = (
            _canonical_total_region(source_region)
            if source_region and source_region != "総数"
            else None
        )
        nationality = source_nationality
        row_kind = "country_or_area"

    return NationalityPopulationTotalRecord(
        period_end=period_end,
        region=region,
        nationality=nationality,
        row_kind=row_kind,
        population=population,
        source_region=source_region,
        source_nationality=source_nationality,
        source_id=source_id,
        source_table="1",
        source_sheet=source_sheet,
        source_row=source_row,
        source_column=source_column,
    )


def _parse_flat_population_totals(
    worksheet,
    *,
    header_row: int,
    columns: Dict[str, int],
    period_end: str,
    source_id: str,
) -> List[NationalityPopulationTotalRecord]:
    records = []
    for source_row, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if not any(value not in (None, "") for value in row):
            continue
        if _compact_label(row[columns["在留資格"]]) != "総数":
            continue
        if "在留目的" in columns and _compact_label(row[columns["在留目的"]]) != "計":
            continue

        source_region = _compact_label(row[columns["州"]])
        source_nationality = _compact_label(row[columns["国籍・地域"]])
        if not source_region or not source_nationality:
            raise SchemaError(
                "Incomplete nationality population labels at source row %d"
                % source_row
            )
        population_column = columns["在留外国人数"]
        records.append(
            _make_population_total_record(
                period_end=period_end,
                source_region=source_region,
                source_nationality=source_nationality,
                population=_total_integer(
                    row[population_column], source_row=source_row
                ),
                source_id=source_id,
                source_sheet=worksheet.title,
                source_row=source_row,
                source_column=population_column + 1,
            )
        )
    return records


def _parse_wide_population_totals(
    worksheet,
    *,
    period_end: str,
    source_id: str,
) -> List[NationalityPopulationTotalRecord]:
    total_column = None
    for row in worksheet.iter_rows(
        min_row=1, max_row=min(6, worksheet.max_row), values_only=True
    ):
        for column, value in enumerate(row):
            if _compact_label(value) == "総数":
                total_column = column
                break
        if total_column is not None:
            break
    if total_column is None or total_column == 0:
        raise SchemaError("Legacy Table 1 total column was not found")

    records = []
    current_source_region = None
    for source_row, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if total_column >= len(row) or row[total_column] in (None, ""):
            continue
        labels = [
            _compact_label(value) for value in row[:total_column] if _compact_label(value)
        ]
        if not labels:
            continue
        source_label = labels[-1]
        try:
            population = _total_integer(row[total_column], source_row=source_row)
        except SchemaError:
            if source_row <= 6:
                continue
            raise

        if source_label == "総数":
            source_region = "総数"
            source_nationality = "総数"
        elif source_label in TOTAL_REGION_LABELS:
            current_source_region = source_label
            source_region = source_label
            source_nationality = "総数"
        elif source_label == "無国籍":
            source_region = "無国籍"
            source_nationality = "総数"
        else:
            if current_source_region is None:
                raise SchemaError(
                    "Country row appears before a region total at source row %d"
                    % source_row
                )
            source_region = current_source_region
            source_nationality = source_label

        records.append(
            _make_population_total_record(
                period_end=period_end,
                source_region=source_region,
                source_nationality=source_nationality,
                population=population,
                source_id=source_id,
                source_sheet=worksheet.title,
                source_row=source_row,
                source_column=total_column + 1,
            )
        )
    return records


def _validate_population_total_sum(
    records: Sequence[NationalityPopulationTotalRecord],
) -> None:
    national = [record for record in records if record.row_kind == "national_total"]
    if len(national) != 1:
        raise SchemaError(
            "Expected one national nationality population total; found %d"
            % len(national)
        )
    country_total = sum(
        record.population
        for record in records
        if record.row_kind == "country_or_area"
    )
    if country_total != national[0].population:
        raise SchemaError(
            "Published country-or-area total %d differs from national total %d"
            % (country_total, national[0].population)
        )


def parse_population_nationality_totals(
    path: Path,
    *,
    source_id: str,
) -> List[NationalityPopulationTotalRecord]:
    """Parse the total column from legacy and flat ISA Table 1 layouts."""

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        _validate_population_total_title(worksheet)
        period_end = _total_period_end(worksheet.title)
        flat_header = _flat_total_header(worksheet)
        if flat_header is None:
            records = _parse_wide_population_totals(
                worksheet,
                period_end=period_end,
                source_id=source_id,
            )
        else:
            header_row, columns = flat_header
            records = _parse_flat_population_totals(
                worksheet,
                header_row=header_row,
                columns=columns,
                period_end=period_end,
                source_id=source_id,
            )
        _validate_population_total_sum(records)
        return records
    finally:
        workbook.close()
