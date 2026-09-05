"""Parsers for NPA context tables and Statistics Bureau population Table 2."""

import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from .errors import SchemaError
from .models import (
    NationalClearanceAnnualRecord,
    OverallPrefectureCrimeRecord,
    PrefecturePopulationRecord,
)


PREFECTURE_BASE_LABELS = (
    "北海道",
    "青森",
    "岩手",
    "宮城",
    "秋田",
    "山形",
    "福島",
    "茨城",
    "栃木",
    "群馬",
    "埼玉",
    "千葉",
    "東京",
    "神奈川",
    "新潟",
    "富山",
    "石川",
    "福井",
    "山梨",
    "長野",
    "岐阜",
    "静岡",
    "愛知",
    "三重",
    "滋賀",
    "京都",
    "大阪",
    "兵庫",
    "奈良",
    "和歌山",
    "鳥取",
    "島根",
    "岡山",
    "広島",
    "山口",
    "徳島",
    "香川",
    "愛媛",
    "高知",
    "福岡",
    "佐賀",
    "長崎",
    "熊本",
    "大分",
    "宮崎",
    "鹿児島",
    "沖縄",
)


def _canonical_prefecture(base: str) -> Optional[str]:
    if base not in PREFECTURE_BASE_LABELS:
        return None
    if base == "北海道":
        return base
    if base == "東京":
        return "東京都"
    if base in {"京都", "大阪"}:
        return base + "府"
    return base + "県"


PREFECTURE_PARENT_REGION: Dict[str, str] = {
    "北海道": "北海道",
    **{
        name: "東北"
        for name in ("青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県")
    },
    "東京都": "東京都",
    **{
        name: "関東"
        for name in (
            "茨城県",
            "栃木県",
            "群馬県",
            "埼玉県",
            "千葉県",
            "神奈川県",
            "新潟県",
            "山梨県",
            "長野県",
            "静岡県",
        )
    },
    **{
        name: "中部"
        for name in ("富山県", "石川県", "福井県", "岐阜県", "愛知県", "三重県")
    },
    **{
        name: "近畿"
        for name in ("滋賀県", "京都府", "大阪府", "兵庫県", "奈良県", "和歌山県")
    },
    **{
        name: "中国"
        for name in ("鳥取県", "島根県", "岡山県", "広島県", "山口県")
    },
    **{name: "四国" for name in ("徳島県", "香川県", "愛媛県", "高知県")},
    **{
        name: "九州"
        for name in (
            "福岡県",
            "佐賀県",
            "長崎県",
            "熊本県",
            "大分県",
            "宮崎県",
            "鹿児島県",
            "沖縄県",
        )
    },
}

POLICE_REGIONS = {"東北", "関東", "中部", "近畿", "中国", "四国", "九州"}
HOKKAIDO_SUBREGIONS = {"札幌", "函館", "旭川", "釧路", "北見"}


def _clean(value: object) -> str:
    return "".join(str(value or "").split())


def _integer(value: object, *, label: str, source_row: int) -> int:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        raise SchemaError("Missing %s at source row %d" % (label, source_row))
    if not numeric.is_integer() or numeric < 0:
        raise SchemaError("Invalid %s at source row %d" % (label, source_row))
    return int(numeric)


def _workbook_title(worksheet, *, max_row: int = 5) -> str:
    return _clean(
        " ".join(
            str(value)
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, values_only=True)
            for value in row
            if value not in (None, "")
        )
    )


def _latest_national_row(worksheet) -> Tuple[int, int]:
    year_rows = _annual_national_rows(worksheet)
    if not year_rows:
        raise SchemaError("Table 3 annual total rows were not found")
    return max(year_rows)


def _annual_national_rows(worksheet) -> List[Tuple[int, int]]:
    year_rows = []
    seen_years = set()
    for source_row in range(1, worksheet.max_row + 1):
        value = worksheet.cell(source_row, 2).value
        match = re.search(r"(?:^|\D)(\d{4})(?:\D|$)", str(value or ""))
        if match and all(
            worksheet.cell(source_row, column).value not in (None, "")
            for column in (3, 5, 6)
        ):
            year = int(match.group(1))
            if year in seen_years:
                raise SchemaError("Duplicate Table 3 annual row for %d" % year)
            seen_years.add(year)
            year_rows.append((year, source_row))
    return sorted(year_rows)


def _criminal_code_total_worksheet(workbook):
    if "刑法犯総数" not in workbook.sheetnames:
        raise SchemaError("Table 3 刑法犯総数 sheet was not found")
    worksheet = workbook["刑法犯総数"]
    title = _workbook_title(worksheet)
    if "年次別都道府県別" not in title or "認知・検挙件数及び検挙人員" not in title:
        raise SchemaError("Table 3 title was not recognized")
    if "刑法犯総数（交通業過を除く）" not in title:
        raise SchemaError("Table 3 offense scope was not recognized")
    return worksheet


def parse_npa_all_person_annual_clearances(
    path: Path,
    *,
    source_id: str,
) -> List[NationalClearanceAnnualRecord]:
    """Parse every published national criminal-code annual total in Table 3."""

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet = _criminal_code_total_worksheet(workbook)
        annual_rows = _annual_national_rows(worksheet)
        if not annual_rows:
            raise SchemaError("Table 3 annual total rows were not found")
        return [
            NationalClearanceAnnualRecord(
                year=year,
                population_scope="all_persons",
                offense_scope="criminal_code_excluding_traffic_negligence",
                geography="日本全国",
                cleared_cases=_integer(
                    worksheet.cell(source_row, 5).value,
                    label="cleared cases",
                    source_row=source_row,
                ),
                cleared_persons=_integer(
                    worksheet.cell(source_row, 6).value,
                    label="cleared persons",
                    source_row=source_row,
                ),
                source_id=source_id,
                source_table="3",
                source_sheet=worksheet.title,
                source_row=source_row,
                source_cases_column=5,
                source_persons_column=6,
            )
            for year, source_row in annual_rows
        ]
    finally:
        workbook.close()


def _crime_geography(label: str) -> Tuple[str, str, Optional[str]]:
    if label in POLICE_REGIONS:
        return label, "police_region", label
    if label in HOKKAIDO_SUBREGIONS:
        return label, "police_subregion", "北海道"
    prefecture = _canonical_prefecture(label)
    if prefecture is None:
        raise SchemaError("Unrecognized Table 3 geography: %s" % label)
    return prefecture, "prefecture", PREFECTURE_PARENT_REGION[prefecture]


def parse_npa_overall_prefecture_crime(
    path: Path,
    *,
    source_id: str,
) -> List[OverallPrefectureCrimeRecord]:
    """Parse latest-year all-person criminal-code counts from NPA Table 3."""

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet = _criminal_code_total_worksheet(workbook)

        year, national_row = _latest_national_row(worksheet)
        records = [
            OverallPrefectureCrimeRecord(
                year=year,
                population_scope="all_persons",
                offense_scope="criminal_code_excluding_traffic_negligence",
                geography="日本",
                geography_type="national",
                parent_region=None,
                geography_semantics="national_aggregate",
                recognized_cases=_integer(
                    worksheet.cell(national_row, 3).value,
                    label="recognized cases",
                    source_row=national_row,
                ),
                cleared_cases=_integer(
                    worksheet.cell(national_row, 5).value,
                    label="cleared cases",
                    source_row=national_row,
                ),
                cleared_persons=_integer(
                    worksheet.cell(national_row, 6).value,
                    label="cleared persons",
                    source_row=national_row,
                ),
                source_id=source_id,
                source_table="3",
                source_sheet=worksheet.title,
                source_row=national_row,
            )
        ]

        for source_row in range(national_row + 1, worksheet.max_row + 1):
            label = _clean(worksheet.cell(source_row, 2).value)
            if not label:
                continue
            if label.startswith("注"):
                break
            if label == "確認用":
                break
            if re.search(r"\d{4}", label):
                continue
            geography, geography_type, parent_region = _crime_geography(label)
            records.append(
                OverallPrefectureCrimeRecord(
                    year=year,
                    population_scope="all_persons",
                    offense_scope="criminal_code_excluding_traffic_negligence",
                    geography=geography,
                    geography_type=geography_type,
                    parent_region=parent_region,
                    geography_semantics="police_reporting_area_unresolved",
                    recognized_cases=_integer(
                        worksheet.cell(source_row, 3).value,
                        label="recognized cases",
                        source_row=source_row,
                    ),
                    cleared_cases=_integer(
                        worksheet.cell(source_row, 5).value,
                        label="cleared cases",
                        source_row=source_row,
                    ),
                    cleared_persons=_integer(
                        worksheet.cell(source_row, 6).value,
                        label="cleared persons",
                        source_row=source_row,
                    ),
                    source_id=source_id,
                    source_table="3",
                    source_sheet=worksheet.title,
                    source_row=source_row,
                )
            )
        return records
    finally:
        workbook.close()


def _population_year_column(worksheet) -> Tuple[int, int]:
    year_columns = []
    for column in range(1, worksheet.max_column + 1):
        match = re.fullmatch(r"\s*(\d{4})年\s*", str(worksheet.cell(5, column).value or ""))
        if match:
            year_columns.append((int(match.group(1)), column))
    if not year_columns:
        raise SchemaError("Table 144 year columns were not found")
    return max(year_columns)


def parse_npa_prefecture_population(
    path: Path,
    *,
    source_id: str,
) -> List[PrefecturePopulationRecord]:
    """Parse latest October 1 total population from NPA Table 144."""

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        title = _workbook_title(worksheet)
        if "144" not in title or "都道府県別人口" not in title:
            raise SchemaError("Table 144 title was not recognized")
        if "1,000人" not in title:
            raise SchemaError("Table 144 population unit was not recognized")

        year, value_column = _population_year_column(worksheet)
        records = []
        for source_row in range(6, worksheet.max_row + 1):
            label = _clean(worksheet.cell(source_row, 3).value)
            if not label:
                continue
            if label.startswith("注"):
                break
            if label == "総人口":
                geography = "日本"
                geography_type = "national"
            else:
                geography = _canonical_prefecture(label)
                if geography is None:
                    raise SchemaError("Unrecognized Table 144 geography: %s" % label)
                geography_type = "prefecture"
            source_value = _integer(
                worksheet.cell(source_row, value_column).value,
                label="population",
                source_row=source_row,
            )
            records.append(
                PrefecturePopulationRecord(
                    year=year,
                    reference_date="%04d-10-01" % year,
                    population_scope="total_population",
                    geography=geography,
                    geography_type=geography_type,
                    parent_region=None,
                    geography_semantics=(
                        "national_aggregate"
                        if geography_type == "national"
                        else "population_estimate_prefecture"
                    ),
                    population=source_value * 1000,
                    source_value=source_value,
                    source_unit="1000_persons",
                    rounding="nearest_1000_persons",
                    source_id=source_id,
                    source_table="144",
                    source_sheet=worksheet.title,
                    source_row=source_row,
                )
            )
        return records
    finally:
        workbook.close()


def parse_statistics_bureau_japanese_population(
    path: Path,
    *,
    source_id: str,
) -> List[PrefecturePopulationRecord]:
    """Parse published Japanese-national population from Statistics Bureau Table 2."""

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        if "第2表" not in workbook.sheetnames:
            raise SchemaError("Statistics Bureau Table 2 title was not recognized")
        worksheet = workbook["第2表"]
        title = _workbook_title(worksheet, max_row=1)
        if (
            "第２表" not in title
            or "都道府県" not in title
            or "総人口、日本人人口" not in title
        ):
            raise SchemaError("Statistics Bureau Table 2 title was not recognized")
        if "単位千人" not in _clean(worksheet.cell(4, 12).value):
            raise SchemaError("Statistics Bureau Table 2 population unit was not recognized")
        if (
            _clean(worksheet.cell(6, 5).value) != "総人口"
            or _clean(worksheet.cell(6, 9).value) != "日本人人口"
            or _clean(worksheet.cell(9, 9).value) != "男女計"
        ):
            raise SchemaError("Statistics Bureau Table 2 headers were not recognized")

        year_match = re.search(r"(\d{4})年10月[1１]日", title)
        if year_match is None:
            raise SchemaError("Statistics Bureau Table 2 reference date was not recognized")
        year = int(year_match.group(1))
        records = []
        for source_row in range(12, worksheet.max_row + 1):
            national_label = _clean(worksheet.cell(source_row, 1).value)
            prefecture_code = _clean(worksheet.cell(source_row, 2).value)
            prefecture_label = _clean(worksheet.cell(source_row, 3).value)
            if national_label == "全国":
                geography = "日本"
                geography_type = "national"
            elif re.fullmatch(r"\d{2}", prefecture_code) and prefecture_label:
                geography = (
                    prefecture_label
                    if prefecture_label in PREFECTURE_PARENT_REGION
                    else _canonical_prefecture(prefecture_label)
                )
                if geography is None:
                    raise SchemaError(
                        "Unrecognized Statistics Bureau Table 2 geography: %s"
                        % prefecture_label
                    )
                geography_type = "prefecture"
            else:
                continue

            source_value = _integer(
                worksheet.cell(source_row, 9).value,
                label="Japanese population",
                source_row=source_row,
            )
            records.append(
                PrefecturePopulationRecord(
                    year=year,
                    reference_date="%04d-10-01" % year,
                    population_scope="japanese_population",
                    geography=geography,
                    geography_type=geography_type,
                    parent_region=None,
                    geography_semantics=(
                        "national_aggregate"
                        if geography_type == "national"
                        else "population_estimate_prefecture"
                    ),
                    population=source_value * 1000,
                    source_value=source_value,
                    source_unit="1000_persons",
                    rounding="nearest_1000_persons",
                    source_id=source_id,
                    source_table="2",
                    source_sheet=worksheet.title.strip(),
                    source_row=source_row,
                )
            )
        if not any(record.geography_type == "national" for record in records):
            raise SchemaError("Statistics Bureau Table 2 national row was not found")
        return records
    finally:
        workbook.close()


INTERCENSAL_POPULATION_SHEETS = {
    "総人口 (2015年～2020年)": "total_population",
    "日本人人口 (2015年～2020年)": "japanese_population",
}


def _intercensal_year_columns(worksheet) -> List[Tuple[int, int]]:
    year_columns = []
    for column in range(1, worksheet.max_column + 1):
        match = re.fullmatch(r"\s*(\d{4})年\s*", str(worksheet.cell(9, column).value or ""))
        if match:
            year_columns.append((int(match.group(1)), column))
    if not year_columns:
        raise SchemaError("Statistics Bureau Table 5 year columns were not found")
    return year_columns


def parse_statistics_bureau_intercensal_population(
    path: Path,
    *,
    source_id: str,
) -> List[PrefecturePopulationRecord]:
    """Parse 2015--2020 adjusted total and Japanese population from Table 5."""

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        missing_sheets = sorted(
            set(INTERCENSAL_POPULATION_SHEETS) - set(workbook.sheetnames)
        )
        if missing_sheets:
            raise SchemaError(
                "Statistics Bureau Table 5 sheets were not found: %s"
                % ", ".join(missing_sheets)
            )

        records = []
        for sheet_name, population_scope in INTERCENSAL_POPULATION_SHEETS.items():
            worksheet = workbook[sheet_name]
            title = _workbook_title(worksheet, max_row=5)
            if (
                "第５表" not in title
                or "都道府県別人口" not in title
                or "各年10月1日現在" not in title
                or "総人口、日本人" not in title
            ):
                raise SchemaError("Statistics Bureau Table 5 title was not recognized")
            if "単位千人" not in _clean(worksheet.cell(5, 1).value):
                raise SchemaError("Statistics Bureau Table 5 population unit was not recognized")

            scope_header = _clean(
                " ".join(
                    str(value)
                    for value in next(
                        worksheet.iter_rows(min_row=7, max_row=7, values_only=True)
                    )
                    if value not in (None, "")
                )
            )
            expected_scope_header = (
                "総人口"
                if population_scope == "total_population"
                else "日本人人口"
            )
            if expected_scope_header not in scope_header:
                raise SchemaError(
                    "Statistics Bureau Table 5 population scope was not recognized"
                )

            year_columns = _intercensal_year_columns(worksheet)
            found_national = False
            for source_row in range(11, worksheet.max_row + 1):
                national_label = _clean(worksheet.cell(source_row, 1).value)
                prefecture_code = _clean(worksheet.cell(source_row, 2).value)
                prefecture_label = _clean(worksheet.cell(source_row, 3).value)
                if national_label == "全国":
                    geography = "日本"
                    geography_type = "national"
                    found_national = True
                elif re.fullmatch(r"\d{2}", prefecture_code) and prefecture_label:
                    geography = (
                        prefecture_label
                        if prefecture_label in PREFECTURE_PARENT_REGION
                        else _canonical_prefecture(prefecture_label)
                    )
                    if geography is None:
                        raise SchemaError(
                            "Unrecognized Statistics Bureau Table 5 geography: %s"
                            % prefecture_label
                        )
                    geography_type = "prefecture"
                else:
                    continue

                for year, value_column in year_columns:
                    source_value = _integer(
                        worksheet.cell(source_row, value_column).value,
                        label="population",
                        source_row=source_row,
                    )
                    records.append(
                        PrefecturePopulationRecord(
                            year=year,
                            reference_date="%04d-10-01" % year,
                            population_scope=population_scope,
                            geography=geography,
                            geography_type=geography_type,
                            parent_region=None,
                            geography_semantics=(
                                "national_aggregate"
                                if geography_type == "national"
                                else "population_estimate_prefecture"
                            ),
                            population=source_value * 1000,
                            source_value=source_value,
                            source_unit="1000_persons",
                            rounding="nearest_1000_persons",
                            source_id=source_id,
                            source_table="5",
                            source_sheet=worksheet.title,
                            source_row=source_row,
                        )
                    )
            if not found_national:
                raise SchemaError(
                    "Statistics Bureau Table 5 national row was not found in %s"
                    % sheet_name
                )
        return records
    finally:
        workbook.close()
