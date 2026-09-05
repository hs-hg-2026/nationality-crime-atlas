"""Parser for NPA tables 130 and 131 national nationality totals."""

import re
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import load_workbook

from .errors import SchemaError
from .models import NationalClearanceAnnualRecord, NationalityCrimeRecord


VALID_TABLES = {
    "130": "all_foreign",
    "131": "visiting_foreign",
}


def _clean_label(value: object) -> Optional[str]:
    if value is None:
        return None
    cleaned = "".join(str(value).split())
    return cleaned or None


def _find_metric_columns(worksheet) -> Tuple[int, int]:
    for row in worksheet.iter_rows(min_row=1, max_row=15, values_only=True):
        cleaned = [_clean_label(value) for value in row]
        for index in range(len(cleaned) - 1):
            if cleaned[index] == "件数" and cleaned[index + 1] == "人員":
                return index, index + 1
    raise SchemaError("Adjacent 件数 and 人員 headers were not found")


def _find_criminal_code_metric_columns(worksheet) -> Tuple[int, int]:
    header_rows = list(
        worksheet.iter_rows(min_row=1, max_row=15, values_only=True)
    )
    max_columns = max(len(row) for row in header_rows)
    for column in range(max_columns - 1):
        column_labels = [
            _clean_label(row[column]) if column < len(row) else None
            for row in header_rows
        ]
        if "刑法犯" not in column_labels or "計" not in column_labels:
            continue
        for row in header_rows:
            if column + 1 >= len(row):
                continue
            if (
                _clean_label(row[column]) == "件数"
                and _clean_label(row[column + 1]) == "人員"
            ):
                return column, column + 1
    raise SchemaError("刑法犯計の adjacent 件数 and 人員 headers were not found")


def _validate_table_title(worksheet, table_id: str) -> None:
    title = " ".join(
        str(value)
        for row in worksheet.iter_rows(min_row=1, max_row=5, values_only=True)
        for value in row
        if value not in (None, "")
    )
    if table_id not in title:
        raise SchemaError("Workbook table title does not match table_id %s" % table_id)


def _latest_year_and_row(worksheet, metric_column: int) -> Tuple[int, int]:
    year_rows = _annual_year_rows(worksheet, metric_column)
    if not year_rows:
        raise SchemaError("Annual rows were not found before the nationality breakdown")
    return max(year_rows)


def _annual_year_rows(worksheet, metric_column: int) -> List[Tuple[int, int]]:
    year_rows = []
    seen_years = set()
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        for value in row[:metric_column]:
            match = re.fullmatch(r"\s*(\d{4})年\s*", str(value or ""))
            if match:
                year = int(match.group(1))
                if year in seen_years:
                    raise SchemaError("Duplicate annual row for %d" % year)
                seen_years.add(year)
                year_rows.append((year, row_index))
                break
    return sorted(year_rows)


def _integer(value: object, source_row: int) -> int:
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        raise SchemaError("Missing numeric total at source row %d" % source_row)
    if not numeric.is_integer():
        raise SchemaError("Non-integer total at source row %d" % source_row)
    return int(numeric)


def parse_npa_nationality_totals(
    path: Path,
    *,
    table_id: str,
    source_id: str,
) -> List[NationalityCrimeRecord]:
    """Parse the latest-year nationality rows from NPA table 130 or 131."""

    if table_id not in VALID_TABLES:
        raise ValueError("table_id must be one of: 130, 131")

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        _validate_table_title(worksheet, table_id)
        cases_column, persons_column = _find_metric_columns(worksheet)
        (
            criminal_code_cases_column,
            criminal_code_persons_column,
        ) = _find_criminal_code_metric_columns(worksheet)
        latest_year, latest_year_row = _latest_year_and_row(worksheet, cases_column)

        records = []
        current_region = None
        current_nationality = None
        for source_row, row in enumerate(
            worksheet.iter_rows(min_row=latest_year_row + 1, values_only=True),
            start=latest_year_row + 1,
        ):
            labels = [_clean_label(value) for value in row[:cases_column]]
            if any(label and label.startswith("注") for label in labels):
                break
            if row[cases_column] in (None, "") and row[persons_column] in (None, ""):
                continue

            primary = labels[1] if len(labels) > 1 else None
            country = labels[2] if len(labels) > 2 else None
            detail = labels[3] if len(labels) > 3 else None

            if primary and "州" in primary:
                current_region = primary
                current_nationality = None
                nationality = None
                subcategory = None
                row_kind = "region_total"
            elif primary in ("無国籍", "国籍不明"):
                current_region = None
                current_nationality = primary
                nationality = primary
                subcategory = None
                row_kind = "country"
            elif country:
                current_nationality = country
                nationality = country
                subcategory = detail
                row_kind = "subcategory" if detail else "country"
            elif primary and detail:
                current_nationality = primary
                nationality = primary
                subcategory = detail
                row_kind = "subcategory"
            elif detail and current_nationality:
                nationality = current_nationality
                subcategory = detail
                row_kind = "subcategory"
            else:
                raise SchemaError("Could not classify nationality row %d" % source_row)

            records.append(
                NationalityCrimeRecord(
                    year=latest_year,
                    population_scope=VALID_TABLES[table_id],
                    region=current_region,
                    nationality=nationality,
                    subcategory=subcategory,
                    row_kind=row_kind,
                    cleared_cases=_integer(row[cases_column], source_row),
                    cleared_persons=_integer(row[persons_column], source_row),
                    criminal_code_cleared_cases=_integer(
                        row[criminal_code_cases_column], source_row
                    ),
                    criminal_code_cleared_persons=_integer(
                        row[criminal_code_persons_column], source_row
                    ),
                    source_id=source_id,
                    source_table=table_id,
                    source_sheet=worksheet.title.strip(),
                    source_row=source_row,
                )
            )
        return records
    finally:
        workbook.close()


def parse_npa_nationality_annual_clearances(
    path: Path,
    *,
    table_id: str,
    source_id: str,
) -> List[NationalClearanceAnnualRecord]:
    """Parse every published national criminal-code annual total in Table 130/131."""

    if table_id not in VALID_TABLES:
        raise ValueError("table_id must be one of: 130, 131")

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        _validate_table_title(worksheet, table_id)
        cases_column, persons_column = _find_criminal_code_metric_columns(worksheet)
        annual_rows = _annual_year_rows(worksheet, cases_column)
        if not annual_rows:
            raise SchemaError("Annual criminal-code totals were not found")

        records = []
        for year, source_row in annual_rows:
            records.append(
                NationalClearanceAnnualRecord(
                    year=year,
                    population_scope=VALID_TABLES[table_id],
                    offense_scope="criminal_code_figure4_basis",
                    geography="日本全国",
                    cleared_cases=_integer(
                        worksheet.cell(source_row, cases_column + 1).value,
                        source_row,
                    ),
                    cleared_persons=_integer(
                        worksheet.cell(source_row, persons_column + 1).value,
                        source_row,
                    ),
                    source_id=source_id,
                    source_table=table_id,
                    source_sheet=worksheet.title.strip(),
                    source_row=source_row,
                    source_cases_column=cases_column + 1,
                    source_persons_column=persons_column + 1,
                )
            )
        return records
    finally:
        workbook.close()
