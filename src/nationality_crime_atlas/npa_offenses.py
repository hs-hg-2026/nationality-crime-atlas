"""Parsers for official NPA top-level offense groups used in compositions."""

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from .errors import SchemaError
from .models import AllPersonOffenseGroupRecord, NationalityOffenseGroupRecord


TOP_LEVEL_CRIMINAL_CODE_GROUP_IDS = (
    "heinous",
    "assaultive",
    "theft",
    "intellectual",
    "morals",
    "other_criminal_code",
)


@dataclass(frozen=True)
class _NationalityGroupSpec:
    offense_id: str
    label: str
    parent_id: Optional[str]
    level: int
    sheet_name: str
    cases_column: int
    header_token: str


NATIONALITY_GROUP_SPECS = (
    _NationalityGroupSpec(
        "all_offenses", "総数", None, 0, "01", 5, "総数"
    ),
    _NationalityGroupSpec(
        "criminal_code", "刑法犯", "all_offenses", 1, "01", 7, "刑法犯"
    ),
    _NationalityGroupSpec(
        "heinous", "凶悪犯", "criminal_code", 2, "01", 9, "凶悪犯"
    ),
    _NationalityGroupSpec(
        "assaultive", "粗暴犯", "criminal_code", 2, "01", 20, "粗暴犯"
    ),
    _NationalityGroupSpec(
        "theft", "窃盗犯", "criminal_code", 2, "02", 9, "窃盗犯"
    ),
    _NationalityGroupSpec(
        "intellectual", "知能犯", "criminal_code", 2, "02", 13, "知能犯"
    ),
    _NationalityGroupSpec(
        "morals", "風俗犯", "criminal_code", 2, "03", 7, "風俗犯"
    ),
    _NationalityGroupSpec(
        "other_criminal_code",
        "その他の刑法犯",
        "criminal_code",
        2,
        "03",
        16,
        "その他の刑法犯",
    ),
    _NationalityGroupSpec(
        "special_law", "特別法犯", "all_offenses", 1, "04", 5, "特別法犯"
    ),
)


@dataclass(frozen=True)
class _AllPersonGroupSpec:
    offense_id: str
    label: str
    parent_id: Optional[str]
    level: int
    sheet_name: str
    header_token: str


ALL_PERSON_GROUP_SPECS = (
    _AllPersonGroupSpec(
        "criminal_code",
        "刑法犯",
        None,
        1,
        "刑法犯総数",
        "刑法犯総数（交通業過を除く）",
    ),
    _AllPersonGroupSpec("heinous", "凶悪犯", "criminal_code", 2, "A", "凶悪犯"),
    _AllPersonGroupSpec(
        "assaultive", "粗暴犯", "criminal_code", 2, "B", "粗暴犯"
    ),
    _AllPersonGroupSpec("theft", "窃盗犯", "criminal_code", 2, "C", "窃盗犯"),
    _AllPersonGroupSpec(
        "intellectual", "知能犯", "criminal_code", 2, "D", "知能犯"
    ),
    _AllPersonGroupSpec("morals", "風俗犯", "criminal_code", 2, "E", "風俗犯"),
    _AllPersonGroupSpec(
        "other_criminal_code",
        "その他の刑法犯",
        "criminal_code",
        2,
        "F",
        "その他の刑法犯",
    ),
)


VALID_TABLES = {"130": "all_foreign", "131": "visiting_foreign"}


def _clean_label(value: object) -> Optional[str]:
    if value is None:
        return None
    cleaned = "".join(str(value).split())
    return cleaned or None


def _severity_role(offense_id: str) -> str:
    if offense_id == "heinous":
        return "official_high_severity_category"
    return "not_a_project_severity_classification"


def _integer(value: object, *, source_sheet: str, source_row: int) -> int:
    try:
        numeric = float(value)
    except (TypeError, ValueError) as error:
        raise SchemaError(
            "Missing numeric offense value at %s row %d"
            % (source_sheet, source_row)
        ) from error
    if not numeric.is_integer() or numeric < 0:
        raise SchemaError(
            "Offense value must be a non-negative integer at %s row %d"
            % (source_sheet, source_row)
        )
    return int(numeric)


def _normalized_worksheets(workbook) -> Dict[str, object]:
    worksheets: Dict[str, object] = {}
    for worksheet in workbook.worksheets:
        name = worksheet.title.strip()
        if name in worksheets:
            raise SchemaError("Workbook has duplicate normalized sheet name %s" % name)
        worksheets[name] = worksheet
    return worksheets


def _validate_table_title(worksheet, table_id: str) -> None:
    title = " ".join(
        str(value)
        for row in worksheet.iter_rows(min_row=1, max_row=5, values_only=True)
        for value in row
        if value not in (None, "")
    )
    if table_id not in title:
        raise SchemaError(
            "Workbook table title does not match table_id %s in sheet %s"
            % (table_id, worksheet.title.strip())
        )


def _validate_header(worksheet, spec: _NationalityGroupSpec) -> None:
    header_values = [
        _clean_label(value)
        for row in worksheet.iter_rows(min_row=4, max_row=8, values_only=True)
        for value in row
    ]
    if _clean_label(spec.header_token) not in header_values:
        raise SchemaError(
            "Required offense header %s was not found in sheet %s"
            % (spec.label, worksheet.title.strip())
        )
    cases_index = spec.cases_column - 1
    for row in worksheet.iter_rows(min_row=4, max_row=9, values_only=True):
        if cases_index + 1 >= len(row):
            continue
        if (
            _clean_label(row[cases_index]) == "件数"
            and _clean_label(row[cases_index + 1]) == "人員"
        ):
            return
    raise SchemaError(
        "Adjacent 件数 and 人員 headers were not found for %s in sheet %s"
        % (spec.label, worksheet.title.strip())
    )


def _latest_year_and_row(worksheet, metric_column: int) -> Tuple[int, int]:
    year_rows = []
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        for value in row[: metric_column - 1]:
            match = re.fullmatch(r"\s*(\d{4})年\s*", str(value or ""))
            if match:
                year_rows.append((int(match.group(1)), row_index))
                break
    if not year_rows:
        raise SchemaError(
            "Annual rows were not found in sheet %s" % worksheet.title.strip()
        )
    return max(year_rows)


def _classify_nationality_rows(
    worksheet,
    *,
    first_row: int,
    cases_column: int,
) -> Iterable[Tuple[int, str, Optional[str], Optional[str], Optional[str], Sequence[object]]]:
    current_region = None
    current_nationality = None
    records_started = False
    for source_row, row in enumerate(
        worksheet.iter_rows(min_row=first_row, values_only=True), start=first_row
    ):
        labels = [_clean_label(value) for value in row[: cases_column - 1]]
        if any(label and label.startswith("注") for label in labels):
            break
        cases_index = cases_column - 1
        if cases_index + 1 >= len(row):
            continue
        if row[cases_index] in (None, "") and row[cases_index + 1] in (None, ""):
            if records_started:
                break
            continue

        primary = labels[1] if len(labels) > 1 else None
        country = labels[2] if len(labels) > 2 else None
        detail = labels[3] if len(labels) > 3 else None
        if primary and "州" in primary:
            current_region = primary
            current_nationality = None
            nationality = None
            subcategory = None
            row_kind = "region_total"
        elif primary in ("無国籍", "国籍不明"):
            current_region = None
            current_nationality = primary
            nationality = primary
            subcategory = None
            row_kind = "country"
        elif country:
            current_nationality = country
            nationality = country
            subcategory = detail
            row_kind = "subcategory" if detail else "country"
        elif primary and detail:
            current_nationality = primary
            nationality = primary
            subcategory = detail
            row_kind = "subcategory"
        elif detail and current_nationality:
            nationality = current_nationality
            subcategory = detail
            row_kind = "subcategory"
        else:
            raise SchemaError(
                "Could not classify nationality row %d in sheet %s"
                % (source_row, worksheet.title.strip())
            )
        records_started = True
        yield (
            source_row,
            row_kind,
            current_region,
            nationality,
            subcategory,
            row,
        )


def _entity_key(record: NationalityOffenseGroupRecord) -> Tuple[object, ...]:
    return (
        record.year,
        record.row_kind,
        record.region,
        record.nationality,
        record.subcategory,
    )


def _validate_nationality_group_totals(
    records: Sequence[NationalityOffenseGroupRecord],
) -> None:
    grouped: Dict[Tuple[object, ...], Dict[str, NationalityOffenseGroupRecord]] = {}
    for record in records:
        grouped.setdefault(_entity_key(record), {})[record.offense_id] = record
    expected_ids = {spec.offense_id for spec in NATIONALITY_GROUP_SPECS}
    for entity, by_offense in grouped.items():
        if set(by_offense) != expected_ids:
            raise SchemaError("Offense groups are incomplete for entity %r" % (entity,))
        criminal = by_offense["criminal_code"]
        all_offenses = by_offense["all_offenses"]
        special = by_offense["special_law"]
        for metric in ("cleared_cases", "cleared_persons"):
            observed = sum(
                getattr(by_offense[offense_id], metric)
                for offense_id in TOP_LEVEL_CRIMINAL_CODE_GROUP_IDS
            )
            if observed != getattr(criminal, metric):
                raise SchemaError(
                    "Top-level groups do not sum to criminal_code for entity %r metric %s"
                    % (entity, metric)
                )
            if getattr(criminal, metric) + getattr(special, metric) != getattr(
                all_offenses, metric
            ):
                raise SchemaError(
                    "Criminal-code and special-law totals do not sum to all offenses "
                    "for entity %r metric %s" % (entity, metric)
                )


def parse_npa_nationality_offense_groups(
    path: Path,
    *,
    table_id: str,
    source_id: str,
) -> List[NationalityOffenseGroupRecord]:
    """Parse official top-level offense groups from NPA Table 130 or 131."""

    if table_id not in VALID_TABLES:
        raise ValueError("table_id must be one of: 130, 131")
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheets = _normalized_worksheets(workbook)
        required_sheets = {spec.sheet_name for spec in NATIONALITY_GROUP_SPECS}
        missing = sorted(required_sheets - set(worksheets))
        if missing:
            raise SchemaError("Workbook is missing required sheet(s): %s" % ", ".join(missing))

        records: List[NationalityOffenseGroupRecord] = []
        entity_sets: Dict[str, set] = {}
        for spec in NATIONALITY_GROUP_SPECS:
            worksheet = worksheets[spec.sheet_name]
            _validate_table_title(worksheet, table_id)
            _validate_header(worksheet, spec)
            year, annual_row = _latest_year_and_row(worksheet, spec.cases_column)
            cases_index = spec.cases_column - 1
            annual_values = next(
                worksheet.iter_rows(
                    min_row=annual_row, max_row=annual_row, values_only=True
                )
            )
            offense_records = [
                NationalityOffenseGroupRecord(
                    year=year,
                    population_scope=VALID_TABLES[table_id],
                    region=None,
                    nationality=None,
                    subcategory=None,
                    row_kind="annual_total",
                    offense_id=spec.offense_id,
                    offense_label=spec.label,
                    offense_parent_id=spec.parent_id,
                    offense_level=spec.level,
                    official_severity_role=_severity_role(spec.offense_id),
                    cleared_cases=_integer(
                        annual_values[cases_index],
                        source_sheet=worksheet.title.strip(),
                        source_row=annual_row,
                    ),
                    cleared_persons=_integer(
                        annual_values[cases_index + 1],
                        source_sheet=worksheet.title.strip(),
                        source_row=annual_row,
                    ),
                    source_id=source_id,
                    source_table=table_id,
                    source_sheet=worksheet.title.strip(),
                    source_row=annual_row,
                    source_cases_column=spec.cases_column,
                    source_persons_column=spec.cases_column + 1,
                )
            ]
            for (
                source_row,
                row_kind,
                region,
                nationality,
                subcategory,
                row,
            ) in _classify_nationality_rows(
                worksheet,
                first_row=annual_row + 1,
                cases_column=spec.cases_column,
            ):
                offense_records.append(
                    NationalityOffenseGroupRecord(
                        year=year,
                        population_scope=VALID_TABLES[table_id],
                        region=region,
                        nationality=nationality,
                        subcategory=subcategory,
                        row_kind=row_kind,
                        offense_id=spec.offense_id,
                        offense_label=spec.label,
                        offense_parent_id=spec.parent_id,
                        offense_level=spec.level,
                        official_severity_role=_severity_role(spec.offense_id),
                        cleared_cases=_integer(
                            row[cases_index],
                            source_sheet=worksheet.title.strip(),
                            source_row=source_row,
                        ),
                        cleared_persons=_integer(
                            row[cases_index + 1],
                            source_sheet=worksheet.title.strip(),
                            source_row=source_row,
                        ),
                        source_id=source_id,
                        source_table=table_id,
                        source_sheet=worksheet.title.strip(),
                        source_row=source_row,
                        source_cases_column=spec.cases_column,
                        source_persons_column=spec.cases_column + 1,
                    )
                )
            entities = {_entity_key(record) for record in offense_records}
            entity_sets[spec.offense_id] = entities
            records.extend(offense_records)

        reference_entities = entity_sets["criminal_code"]
        for offense_id, entities in entity_sets.items():
            if entities != reference_entities:
                raise SchemaError(
                    "Nationality entities differ between criminal_code and %s"
                    % offense_id
                )
        _validate_nationality_group_totals(records)
        return records
    finally:
        workbook.close()


def _latest_all_person_row(worksheet) -> Tuple[int, int]:
    rows = []
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        label = str(row[1] or "") if len(row) > 1 else ""
        match = re.match(r"\s*(\d{4})(?:\D|$)", label)
        if match:
            rows.append((int(match.group(1)), row_index))
    if not rows:
        raise SchemaError(
            "Annual all-person rows were not found in sheet %s"
            % worksheet.title.strip()
        )
    return max(rows)


def _validate_all_person_totals(records: Sequence[AllPersonOffenseGroupRecord]) -> None:
    by_offense = {record.offense_id: record for record in records}
    expected_ids = {spec.offense_id for spec in ALL_PERSON_GROUP_SPECS}
    if set(by_offense) != expected_ids:
        raise SchemaError("All-person offense groups are incomplete")
    for metric in ("recognized_cases", "cleared_cases", "cleared_persons"):
        observed = sum(
            getattr(by_offense[offense_id], metric)
            for offense_id in TOP_LEVEL_CRIMINAL_CODE_GROUP_IDS
        )
        if observed != getattr(by_offense["criminal_code"], metric):
            raise SchemaError(
                "All-person top-level groups do not sum to criminal_code for %s"
                % metric
            )


def parse_npa_all_person_offense_groups(
    path: Path, *, source_id: str
) -> List[AllPersonOffenseGroupRecord]:
    """Parse nationwide all-person totals for the six criminal-code groups."""

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheets = {worksheet.title.strip(): worksheet for worksheet in workbook.worksheets}
        missing = sorted(
            {spec.sheet_name for spec in ALL_PERSON_GROUP_SPECS} - set(worksheets)
        )
        if missing:
            raise SchemaError("Workbook is missing required sheet(s): %s" % ", ".join(missing))
        records = []
        for spec in ALL_PERSON_GROUP_SPECS:
            worksheet = worksheets[spec.sheet_name]
            title = _clean_label(worksheet.cell(row=4, column=3).value)
            if _clean_label(spec.header_token) not in str(title or ""):
                raise SchemaError(
                    "Required all-person offense header %s was not found in sheet %s"
                    % (spec.label, spec.sheet_name)
                )
            year, source_row = _latest_all_person_row(worksheet)
            values = next(
                worksheet.iter_rows(
                    min_row=source_row, max_row=source_row, values_only=True
                )
            )
            records.append(
                AllPersonOffenseGroupRecord(
                    year=year,
                    population_scope="all_persons",
                    geography="日本全国",
                    offense_id=spec.offense_id,
                    offense_label=spec.label,
                    offense_parent_id=spec.parent_id,
                    offense_level=spec.level,
                    official_severity_role=_severity_role(spec.offense_id),
                    recognized_cases=_integer(
                        values[2], source_sheet=spec.sheet_name, source_row=source_row
                    ),
                    cleared_cases=_integer(
                        values[4], source_sheet=spec.sheet_name, source_row=source_row
                    ),
                    cleared_persons=_integer(
                        values[5], source_sheet=spec.sheet_name, source_row=source_row
                    ),
                    source_id=source_id,
                    source_table="3",
                    source_sheet=spec.sheet_name,
                    source_row=source_row,
                )
            )
        years = {record.year for record in records}
        if len(years) != 1:
            raise SchemaError("All-person offense sheets have inconsistent latest years")
        _validate_all_person_totals(records)
        return records
    finally:
        workbook.close()
