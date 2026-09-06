"""Build annual clearance reference ratios against published populations."""

import csv
import json
import os
import re
import shutil
import tempfile
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Mapping, Optional, Sequence, Tuple

from openpyxl import load_workbook

from .errors import IntegrityError, SchemaError
from .models import (
    NationalClearanceAnnualRecord,
    NationalityPopulationTotalRecord,
    PrefecturePopulationRecord,
)
from .npa_all_residents import (
    parse_npa_all_person_annual_clearances,
    parse_statistics_bureau_intercensal_population,
    parse_statistics_bureau_japanese_population,
)
from .npa_nationality import parse_npa_nationality_annual_clearances
from .population import parse_population_nationality_totals
from .provenance import sha256_file


CLEARANCE_POPULATION_TREND_SCHEMA_VERSION = 1
CONTRACT_SCHEMA_VERSION = 1
LATEST_SCHEMA_VERSION = 1
SUPPORTED_METRICS = ("cleared_cases", "cleared_persons")
SUPPORTED_GROUPS = ("japanese_etc_residual", "all_foreign")
INTERPRETATION_POLICY = "public_data_reference_ratio_not_probability"
MISSING_FOREIGN_POPULATION_REASON = (
    "resident_foreigner_population_source_not_registered_for_year"
)


@dataclass(frozen=True)
class ClearancePopulationTrendContract:
    """Reviewed sources and display semantics for population reference ratios."""

    trend_id: str
    label_ja: str
    label_en: str
    years: Tuple[int, ...]
    all_person_source_id: str
    all_foreign_source_id: str
    japanese_population_sources: Mapping[int, str]
    resident_foreign_population_sources: Mapping[int, str]
    metrics: Tuple[str, ...]
    display_multiplier: int
    display_unit_label_ja: str
    interpretation_policy: str
    ui_caveat: str


@dataclass(frozen=True)
class ClearancePopulationTrendRecord:
    """One annual numerator and population pairing, including explicit refusals."""

    clearance_population_trend_schema_version: int
    trend_id: str
    label_ja: str
    label_en: str
    interpretation_policy: str
    ui_caveat: str
    year: int
    population_group: str
    population_group_label_ja: str
    metric: str
    metric_label_ja: str
    numerator_value: int
    denominator_value: Optional[int]
    quotient: Optional[float]
    display_multiplier: int
    display_unit_label_ja: str
    display_value: Optional[float]
    calculation_status: str
    refusal_reason: Optional[str]
    numerator_source_ids: Tuple[str, ...]
    denominator_source_id: Optional[str]
    population_reference_date: Optional[str]
    population_scope: str
    denominator_rounding: Optional[str]
    derivation_method: str
    derivation_formula: Optional[str]
    source_components: Tuple[Mapping[str, object], ...]
    mismatch_flags: Tuple[str, ...]

    def to_dict(self) -> Dict[str, object]:
        return asdict(self)


@dataclass(frozen=True)
class ClearancePopulationTrendReport:
    """Locations and counts for one immutable population-reference run."""

    output_dir: Path
    jsonl_path: Path
    csv_path: Path
    summary_path: Path
    latest_path: Path
    record_count: int
    calculated_count: int
    refused_count: int


@dataclass(frozen=True)
class _SourceInput:
    source_id: str
    catalog_row: Mapping[str, object]
    raw_path: Path
    raw_sha256: str


@dataclass(frozen=True)
class _JapanesePopulationPoint:
    record: PrefecturePopulationRecord
    source_column: int


def _read_json_object(path: Path, label: str) -> Mapping[str, object]:
    try:
        value = json.loads(Path(path).read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError, UnicodeDecodeError) as error:
        raise SchemaError("Invalid %s JSON: %s" % (label, path)) from error
    if not isinstance(value, dict):
        raise SchemaError("%s JSON must contain an object" % label)
    return value


def _require_mapping(value: object, label: str) -> Mapping[str, object]:
    if not isinstance(value, dict):
        raise SchemaError("%s must be an object" % label)
    return value


def _require_list(value: object, label: str) -> Sequence[object]:
    if not isinstance(value, list):
        raise SchemaError("%s must be an array" % label)
    return value


def _require_string(value: object, label: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise SchemaError("%s must be a non-empty string" % label)
    return value


def _require_int(value: object, label: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int):
        raise SchemaError("%s must be an integer" % label)
    return value


def _require_sha256(value: object, label: str) -> str:
    result = _require_string(value, label)
    if not re.fullmatch(r"[0-9a-f]{64}", result):
        raise SchemaError("%s must be a lowercase SHA-256 digest" % label)
    return result


def _year_source_mapping(value: object, label: str) -> Mapping[int, str]:
    raw = _require_mapping(value, label)
    result = {}
    for raw_year, raw_source_id in raw.items():
        if not isinstance(raw_year, str) or not re.fullmatch(r"\d{4}", raw_year):
            raise SchemaError("%s keys must be four-digit years" % label)
        year = int(raw_year)
        if year in result:
            raise SchemaError("Duplicate year in %s: %d" % (label, year))
        result[year] = _require_string(raw_source_id, "%s[%s]" % (label, raw_year))
    return result


def _load_contract(
    path: Path,
) -> Tuple[ClearancePopulationTrendContract, Mapping[str, str]]:
    data = _read_json_object(path, "clearance population trend contract")
    if data.get("schema_version") != CONTRACT_SCHEMA_VERSION:
        raise SchemaError(
            "Unsupported clearance population trend contract schema_version"
        )
    raw_pins = _require_mapping(data.get("artifact_pins"), "artifact_pins")
    pins = {
        _require_string(source_id, "artifact pin source_id"): _require_sha256(
            digest, "artifact_pins[%s]" % source_id
        )
        for source_id, digest in raw_pins.items()
    }
    item = _require_mapping(data.get("trend"), "trend")
    years = tuple(
        _require_int(year, "years")
        for year in _require_list(item.get("years"), "years")
    )
    if not years or tuple(sorted(set(years))) != years:
        raise SchemaError("years must be a non-empty ascending unique array")
    japanese_sources = _year_source_mapping(
        item.get("japanese_population_sources"),
        "japanese_population_sources",
    )
    foreign_sources = _year_source_mapping(
        item.get("resident_foreign_population_sources"),
        "resident_foreign_population_sources",
    )
    if set(japanese_sources) != set(years):
        raise SchemaError("Japanese population sources must cover every trend year")
    if not set(foreign_sources).issubset(years):
        raise SchemaError("Resident-foreign population source year is outside trend years")
    metrics = tuple(
        _require_string(metric, "metrics")
        for metric in _require_list(item.get("metrics"), "metrics")
    )
    if metrics != SUPPORTED_METRICS:
        raise SchemaError("metrics must contain cleared_cases then cleared_persons")
    display_multiplier = _require_int(
        item.get("display_multiplier"), "display_multiplier"
    )
    if display_multiplier != 1000:
        raise SchemaError("display_multiplier must be 1000")
    interpretation_policy = _require_string(
        item.get("interpretation_policy"), "interpretation_policy"
    )
    if interpretation_policy != INTERPRETATION_POLICY:
        raise SchemaError("Unsupported interpretation_policy")
    all_person_source_id = _require_string(
        item.get("all_person_source_id"), "all_person_source_id"
    )
    all_foreign_source_id = _require_string(
        item.get("all_foreign_source_id"), "all_foreign_source_id"
    )
    if all_person_source_id == all_foreign_source_id:
        raise SchemaError("All-person and all-foreign sources must differ")
    required_sources = {
        all_person_source_id,
        all_foreign_source_id,
        *japanese_sources.values(),
        *foreign_sources.values(),
    }
    if set(pins) != required_sources:
        raise SchemaError("artifact_pins must exactly match trend sources")
    return (
        ClearancePopulationTrendContract(
            trend_id=_require_string(item.get("trend_id"), "trend_id"),
            label_ja=_require_string(item.get("label_ja"), "label_ja"),
            label_en=_require_string(item.get("label_en"), "label_en"),
            years=years,
            all_person_source_id=all_person_source_id,
            all_foreign_source_id=all_foreign_source_id,
            japanese_population_sources=japanese_sources,
            resident_foreign_population_sources=foreign_sources,
            metrics=metrics,
            display_multiplier=display_multiplier,
            display_unit_label_ja=_require_string(
                item.get("display_unit_label_ja"), "display_unit_label_ja"
            ),
            interpretation_policy=interpretation_policy,
            ui_caveat=_require_string(item.get("ui_caveat"), "ui_caveat"),
        ),
        pins,
    )


def _read_catalog(path: Path) -> List[Mapping[str, object]]:
    rows = []
    try:
        with Path(path).open(encoding="utf-8") as handle:
            for line_number, line in enumerate(handle, start=1):
                row = json.loads(line)
                if not isinstance(row, dict):
                    raise SchemaError(
                        "Catalog row must be an object at line %d" % line_number
                    )
                rows.append(row)
    except (FileNotFoundError, json.JSONDecodeError, UnicodeDecodeError) as error:
        raise SchemaError("Invalid artifact catalog: %s" % path) from error
    if not rows:
        raise SchemaError("Artifact catalog is empty")
    return rows


def _safe_join(root: Path, value: object, label: str) -> Path:
    relative = Path(_require_string(value, label))
    if relative.is_absolute() or ".." in relative.parts:
        raise SchemaError("Unsafe %s: %s" % (label, relative))
    return Path(root) / relative


def _select_source_input(
    *,
    catalog_rows: Sequence[Mapping[str, object]],
    source_id: str,
    artifact_pin: str,
    raw_root: Path,
) -> _SourceInput:
    source_rows = [row for row in catalog_rows if row.get("source_id") == source_id]
    matching = [row for row in source_rows if row.get("sha256") == artifact_pin]
    if len(matching) != 1:
        raise IntegrityError(
            "Catalog artifact pin selected %d rows for %s"
            % (len(matching), source_id)
        )
    row = matching[0]
    if row.get("processing_status") != "validated":
        raise SchemaError(
            "Clearance population trend requires validated source %s" % source_id
        )
    raw_path = _safe_join(raw_root, row.get("raw_relpath"), "raw_relpath")
    if not raw_path.is_file():
        raise SchemaError("Raw clearance population input is missing: %s" % raw_path)
    observed_hash = sha256_file(raw_path)
    if observed_hash != artifact_pin:
        raise IntegrityError("Raw artifact differs from artifact pin for %s" % source_id)
    return _SourceInput(
        source_id=source_id,
        catalog_row=row,
        raw_path=raw_path,
        raw_sha256=observed_hash,
    )


def _source_artifact(source: _SourceInput) -> Mapping[str, object]:
    fields = (
        "series_id",
        "dataset",
        "publisher",
        "source_table",
        "source_period",
        "sha256",
        "landing_url",
        "download_url",
        "retrieved_at",
        "revision",
        "verification_level",
    )
    artifact = {}
    for field in fields:
        value = source.catalog_row.get(field)
        if value is None:
            raise SchemaError(
                "Catalog provenance field %s is missing for %s"
                % (field, source.source_id)
            )
        artifact[field] = value
    return artifact


def _clearances_by_year(
    records: Sequence[NationalClearanceAnnualRecord],
    *,
    expected_years: Sequence[int],
    source_id: str,
) -> Mapping[int, NationalClearanceAnnualRecord]:
    indexed = {record.year: record for record in records}
    if len(indexed) != len(records):
        raise SchemaError("Duplicate annual clearance row for %s" % source_id)
    if tuple(sorted(indexed)) != tuple(expected_years):
        raise SchemaError(
            "Annual years differ for %s: expected %r observed %r"
            % (source_id, tuple(expected_years), tuple(sorted(indexed)))
        )
    return indexed


def _intercensal_population_columns(path: Path) -> Mapping[Tuple[str, int], int]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        result = {}
        for worksheet in workbook.worksheets:
            for column in range(1, worksheet.max_column + 1):
                match = re.fullmatch(
                    r"\s*(\d{4})年\s*",
                    str(worksheet.cell(9, column).value or ""),
                )
                if match:
                    result[(worksheet.title, int(match.group(1)))] = column
        return result
    finally:
        workbook.close()


def _japanese_population_points(
    *,
    inputs: Mapping[str, _SourceInput],
    source_by_year: Mapping[int, str],
) -> Mapping[int, _JapanesePopulationPoint]:
    years_by_source = defaultdict(list)
    for year, source_id in source_by_year.items():
        years_by_source[source_id].append(year)

    result = {}
    for source_id, expected_years in years_by_source.items():
        source = inputs[source_id]
        source_table = source.catalog_row.get("source_table")
        if source_table == "5":
            parsed = parse_statistics_bureau_intercensal_population(
                source.raw_path,
                source_id=source_id,
            )
            source_columns = _intercensal_population_columns(source.raw_path)
        elif source_table == "2":
            parsed = parse_statistics_bureau_japanese_population(
                source.raw_path,
                source_id=source_id,
            )
            source_columns = {}
        else:
            raise SchemaError(
                "Unsupported Japanese population source table for %s" % source_id
            )
        national = {
            record.year: record
            for record in parsed
            if record.geography == "日本"
            and record.geography_type == "national"
            and record.population_scope == "japanese_population"
        }
        for year in expected_years:
            record = national.get(year)
            if record is None:
                raise SchemaError(
                    "Japanese population year %d was not found in %s"
                    % (year, source_id)
                )
            source_column = (
                9
                if source_table == "2"
                else source_columns.get((record.source_sheet, year))
            )
            if not isinstance(source_column, int):
                raise SchemaError(
                    "Japanese population source column was not found for %d" % year
                )
            result[year] = _JapanesePopulationPoint(
                record=record,
                source_column=source_column,
            )
    if set(result) != set(source_by_year):
        raise SchemaError("Japanese population year grid is incomplete")
    return result


def _resident_foreign_population_points(
    *,
    inputs: Mapping[str, _SourceInput],
    source_by_year: Mapping[int, str],
) -> Mapping[int, NationalityPopulationTotalRecord]:
    result = {}
    for year, source_id in source_by_year.items():
        parsed = parse_population_nationality_totals(
            inputs[source_id].raw_path,
            source_id=source_id,
        )
        national = [record for record in parsed if record.row_kind == "national_total"]
        if len(national) != 1:
            raise SchemaError(
                "Expected one resident-foreign national total for %s" % source_id
            )
        record = national[0]
        if record.period_end != "%04d-12-31" % year:
            raise SchemaError(
                "Resident-foreign population year %d differs in %s"
                % (year, source_id)
            )
        result[year] = record
    return result


def _clearance_component(
    record: NationalClearanceAnnualRecord,
    metric: str,
    *,
    role: str,
) -> Mapping[str, object]:
    return {
        "source_id": record.source_id,
        "source_table": record.source_table,
        "source_sheet": record.source_sheet,
        "source_row": record.source_row,
        "source_column": (
            record.source_cases_column
            if metric == "cleared_cases"
            else record.source_persons_column
        ),
        "metric": metric,
        "value": getattr(record, metric),
        "role": role,
    }


def _japanese_population_component(
    point: _JapanesePopulationPoint,
) -> Mapping[str, object]:
    record = point.record
    return {
        "source_id": record.source_id,
        "source_table": record.source_table,
        "source_sheet": record.source_sheet,
        "source_row": record.source_row,
        "source_column": point.source_column,
        "metric": "population",
        "value": record.population,
        "published_value": record.source_value,
        "published_unit": record.source_unit,
        "role": "denominator",
    }


def _foreign_population_component(
    record: NationalityPopulationTotalRecord,
) -> Mapping[str, object]:
    return {
        "source_id": record.source_id,
        "source_table": record.source_table,
        "source_sheet": record.source_sheet,
        "source_row": record.source_row,
        "source_column": record.source_column,
        "metric": "population",
        "value": record.population,
        "published_value": record.population,
        "published_unit": "persons",
        "role": "denominator",
    }


def _records(
    contract: ClearancePopulationTrendContract,
    *,
    all_person: Mapping[int, NationalClearanceAnnualRecord],
    all_foreign: Mapping[int, NationalClearanceAnnualRecord],
    japanese_population: Mapping[int, _JapanesePopulationPoint],
    resident_foreign_population: Mapping[int, NationalityPopulationTotalRecord],
) -> List[ClearancePopulationTrendRecord]:
    result = []
    metric_labels = {"cleared_cases": "検挙件数", "cleared_persons": "検挙人員"}
    common_flags = (
        "annual_clearance_flow_vs_point_in_time_population_stock",
        "numerator_residency_scope_not_established",
        "public_data_reference_ratio_not_official_crime_rate",
    )
    for year in contract.years:
        all_person_record = all_person[year]
        all_foreign_record = all_foreign[year]
        japanese_population_point = japanese_population[year]
        if japanese_population_point.record.population <= 0:
            raise SchemaError("Japanese population must be positive for %d" % year)
        for metric in contract.metrics:
            all_person_value = getattr(all_person_record, metric)
            all_foreign_value = getattr(all_foreign_record, metric)
            if all_foreign_value > all_person_value:
                raise SchemaError(
                    "All-foreign clearances exceed all-person clearances for %d %s"
                    % (year, metric)
                )
            japanese_numerator = all_person_value - all_foreign_value
            japanese_denominator = japanese_population_point.record.population
            japanese_quotient = japanese_numerator / japanese_denominator
            japanese_population_source = japanese_population_point.record.source_id
            result.append(
                ClearancePopulationTrendRecord(
                    clearance_population_trend_schema_version=(
                        CLEARANCE_POPULATION_TREND_SCHEMA_VERSION
                    ),
                    trend_id=contract.trend_id,
                    label_ja=contract.label_ja,
                    label_en=contract.label_en,
                    interpretation_policy=contract.interpretation_policy,
                    ui_caveat=contract.ui_caveat,
                    year=year,
                    population_group="japanese_etc_residual",
                    population_group_label_ja=(
                        "日本人等（全国総数−外国人全体の残差）"
                    ),
                    metric=metric,
                    metric_label_ja=metric_labels[metric],
                    numerator_value=japanese_numerator,
                    denominator_value=japanese_denominator,
                    quotient=japanese_quotient,
                    display_multiplier=contract.display_multiplier,
                    display_unit_label_ja=contract.display_unit_label_ja,
                    display_value=japanese_quotient * contract.display_multiplier,
                    calculation_status="calculated",
                    refusal_reason=None,
                    numerator_source_ids=(
                        all_person_record.source_id,
                        all_foreign_record.source_id,
                    ),
                    denominator_source_id=japanese_population_source,
                    population_reference_date=(
                        japanese_population_point.record.reference_date
                    ),
                    population_scope="japanese_population",
                    denominator_rounding=japanese_population_point.record.rounding,
                    derivation_method=(
                        "arithmetic_residual_all_person_minus_all_foreign_division"
                    ),
                    derivation_formula=(
                        "(%s.%s - %s.%s) / %s.population * %d"
                        % (
                            all_person_record.source_id,
                            metric,
                            all_foreign_record.source_id,
                            metric,
                            japanese_population_source,
                            contract.display_multiplier,
                        )
                    ),
                    source_components=(
                        _clearance_component(
                            all_person_record,
                            metric,
                            role="numerator_minuend",
                        ),
                        _clearance_component(
                            all_foreign_record,
                            metric,
                            role="numerator_subtrahend",
                        ),
                        _japanese_population_component(japanese_population_point),
                    ),
                    mismatch_flags=tuple(
                        sorted(
                            (
                                *common_flags,
                                "japanese_numerator_is_arithmetic_residual",
                                "japanese_population_rounded_to_nearest_1000",
                                "october_1_population_reference_date",
                            )
                        )
                    ),
                )
            )

            foreign_population = resident_foreign_population.get(year)
            foreign_common = (
                *common_flags,
                "all_foreign_numerator_vs_resident_foreigner_denominator",
                "december_31_population_reference_date",
            )
            numerator_component = _clearance_component(
                all_foreign_record,
                metric,
                role="numerator",
            )
            if foreign_population is None:
                result.append(
                    ClearancePopulationTrendRecord(
                        clearance_population_trend_schema_version=(
                            CLEARANCE_POPULATION_TREND_SCHEMA_VERSION
                        ),
                        trend_id=contract.trend_id,
                        label_ja=contract.label_ja,
                        label_en=contract.label_en,
                        interpretation_policy=contract.interpretation_policy,
                        ui_caveat=contract.ui_caveat,
                        year=year,
                        population_group="all_foreign",
                        population_group_label_ja=(
                            "外国人全体（分母は在留外国人数）"
                        ),
                        metric=metric,
                        metric_label_ja=metric_labels[metric],
                        numerator_value=all_foreign_value,
                        denominator_value=None,
                        quotient=None,
                        display_multiplier=contract.display_multiplier,
                        display_unit_label_ja=contract.display_unit_label_ja,
                        display_value=None,
                        calculation_status="refused",
                        refusal_reason=MISSING_FOREIGN_POPULATION_REASON,
                        numerator_source_ids=(all_foreign_record.source_id,),
                        denominator_source_id=None,
                        population_reference_date=None,
                        population_scope="resident_foreigner_population",
                        denominator_rounding=None,
                        derivation_method="direct_published_count_division_refused",
                        derivation_formula=None,
                        source_components=(numerator_component,),
                        mismatch_flags=tuple(
                            sorted(
                                (
                                    *foreign_common,
                                    "population_denominator_unavailable",
                                )
                            )
                        ),
                    )
                )
                continue
            if foreign_population.population <= 0:
                raise SchemaError(
                    "Resident-foreign population must be positive for %d" % year
                )
            foreign_quotient = all_foreign_value / foreign_population.population
            result.append(
                ClearancePopulationTrendRecord(
                    clearance_population_trend_schema_version=(
                        CLEARANCE_POPULATION_TREND_SCHEMA_VERSION
                    ),
                    trend_id=contract.trend_id,
                    label_ja=contract.label_ja,
                    label_en=contract.label_en,
                    interpretation_policy=contract.interpretation_policy,
                    ui_caveat=contract.ui_caveat,
                    year=year,
                    population_group="all_foreign",
                    population_group_label_ja=(
                        "外国人全体（分母は在留外国人数）"
                    ),
                    metric=metric,
                    metric_label_ja=metric_labels[metric],
                    numerator_value=all_foreign_value,
                    denominator_value=foreign_population.population,
                    quotient=foreign_quotient,
                    display_multiplier=contract.display_multiplier,
                    display_unit_label_ja=contract.display_unit_label_ja,
                    display_value=foreign_quotient * contract.display_multiplier,
                    calculation_status="calculated",
                    refusal_reason=None,
                    numerator_source_ids=(all_foreign_record.source_id,),
                    denominator_source_id=foreign_population.source_id,
                    population_reference_date=foreign_population.period_end,
                    population_scope="resident_foreigner_population",
                    denominator_rounding="as_published_persons",
                    derivation_method="direct_published_count_division",
                    derivation_formula=(
                        "%s.%s / %s.population * %d"
                        % (
                            all_foreign_record.source_id,
                            metric,
                            foreign_population.source_id,
                            contract.display_multiplier,
                        )
                    ),
                    source_components=(
                        numerator_component,
                        _foreign_population_component(foreign_population),
                    ),
                    mismatch_flags=tuple(sorted(foreign_common)),
                )
            )
    return result


def _contract_timestamp(value: str) -> str:
    normalized = value[:-1] + "+00:00" if value.endswith("Z") else value
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError as error:
        raise SchemaError("generated_at must be an ISO-8601 datetime") from error
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise SchemaError("generated_at must include a timezone offset")
    return parsed.strftime("%Y%m%d_%H%M%S")


def _csv_value(value: object) -> object:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if value is None:
        return ""
    return value


def _write_json(path: Path, value: Mapping[str, object]) -> None:
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _atomic_write_json(path: Path, value: Mapping[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=".%s." % path.name,
        suffix=".tmp",
        dir=path.parent,
    )
    temporary = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
            json.dump(value, handle, ensure_ascii=False, indent=2, sort_keys=True)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        temporary.replace(path)
    finally:
        if temporary.exists():
            temporary.unlink()


def generate_clearance_population_trend(
    *,
    catalog_path: Path,
    raw_root: Path,
    contract_path: Path,
    output_root: Path,
    generated_at: str,
) -> ClearancePopulationTrendReport:
    """Generate one immutable annual clearance-to-population reference dataset."""

    contract, artifact_pins = _load_contract(contract_path)
    catalog_rows = _read_catalog(catalog_path)
    inputs = {
        source_id: _select_source_input(
            catalog_rows=catalog_rows,
            source_id=source_id,
            artifact_pin=artifact_pin,
            raw_root=raw_root,
        )
        for source_id, artifact_pin in artifact_pins.items()
    }
    all_person = _clearances_by_year(
        parse_npa_all_person_annual_clearances(
            inputs[contract.all_person_source_id].raw_path,
            source_id=contract.all_person_source_id,
        ),
        expected_years=contract.years,
        source_id=contract.all_person_source_id,
    )
    all_foreign = _clearances_by_year(
        parse_npa_nationality_annual_clearances(
            inputs[contract.all_foreign_source_id].raw_path,
            table_id="130",
            source_id=contract.all_foreign_source_id,
        ),
        expected_years=contract.years,
        source_id=contract.all_foreign_source_id,
    )
    japanese_population = _japanese_population_points(
        inputs=inputs,
        source_by_year=contract.japanese_population_sources,
    )
    resident_foreign_population = _resident_foreign_population_points(
        inputs=inputs,
        source_by_year=contract.resident_foreign_population_sources,
    )
    records = _records(
        contract,
        all_person=all_person,
        all_foreign=all_foreign,
        japanese_population=japanese_population,
        resident_foreign_population=resident_foreign_population,
    )

    status_counts = Counter(record.calculation_status for record in records)
    destination_root = Path(output_root)
    destination_root.mkdir(parents=True, exist_ok=True)
    destination = destination_root / (
        _contract_timestamp(generated_at) + "_clearance_population_trend"
    )
    if destination.exists():
        raise IntegrityError(
            "Timestamped clearance population trend already exists and was not "
            "overwritten: %s" % destination
        )
    staging = Path(
        tempfile.mkdtemp(prefix=".clearance-population-", dir=destination_root)
    )
    try:
        jsonl_path = staging / "clearance_population_records.jsonl"
        csv_path = staging / "clearance_population_records.csv"
        summary_path = staging / "summary.json"
        dictionaries = [record.to_dict() for record in records]
        with jsonl_path.open("w", encoding="utf-8") as handle:
            for row in dictionaries:
                handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")
        with csv_path.open("w", encoding="utf-8", newline="") as handle:
            fieldnames = tuple(dictionaries[0])
            writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
            writer.writeheader()
            for row in dictionaries:
                writer.writerow({key: _csv_value(value) for key, value in row.items()})
        _write_json(
            summary_path,
            {
                "clearance_population_trend_schema_version": (
                    CLEARANCE_POPULATION_TREND_SCHEMA_VERSION
                ),
                "generated_at": generated_at,
                "trend_id": contract.trend_id,
                "record_count": len(records),
                "year_count": len(contract.years),
                "years": list(contract.years),
                "group_counts": dict(
                    sorted(
                        Counter(record.population_group for record in records).items()
                    )
                ),
                "metric_counts": dict(
                    sorted(Counter(record.metric for record in records).items())
                ),
                "status_counts": dict(sorted(status_counts.items())),
                "artifact_pins": dict(sorted(artifact_pins.items())),
                "source_artifacts": {
                    source_id: _source_artifact(source)
                    for source_id, source in sorted(inputs.items())
                },
                "clearance_population_records_sha256": sha256_file(jsonl_path),
                "clearance_population_records_csv_sha256": sha256_file(csv_path),
            },
        )
        staging.rename(destination)
    finally:
        if staging.exists():
            shutil.rmtree(staging)

    final_jsonl = destination / "clearance_population_records.jsonl"
    final_csv = destination / "clearance_population_records.csv"
    final_summary = destination / "summary.json"
    latest_path = destination_root / "latest.json"
    _atomic_write_json(
        latest_path,
        {
            "clearance_population_trend_schema_version": LATEST_SCHEMA_VERSION,
            "generated_at": generated_at,
            "run_relpath": destination.name,
            "summary_sha256": sha256_file(final_summary),
            "clearance_population_records_sha256": sha256_file(final_jsonl),
            "clearance_population_records_csv_sha256": sha256_file(final_csv),
        },
    )
    return ClearancePopulationTrendReport(
        output_dir=destination,
        jsonl_path=final_jsonl,
        csv_path=final_csv,
        summary_path=final_summary,
        latest_path=latest_path,
        record_count=len(records),
        calculated_count=status_counts.get("calculated", 0),
        refused_count=status_counts.get("refused", 0),
    )
