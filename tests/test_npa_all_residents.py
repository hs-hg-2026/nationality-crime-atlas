import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from nationality_crime_atlas.errors import SchemaError
from nationality_crime_atlas.cli import main as ingest_main
from nationality_crime_atlas.npa_all_residents import (
    parse_npa_overall_prefecture_crime,
    parse_npa_prefecture_population,
    parse_statistics_bureau_japanese_population,
)
from nationality_crime_atlas.pipeline import run_offline_pipeline


def _crime_fixture(path: Path, *, valid_title: bool = True) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "刑法犯総数"
    worksheet["B2"] = (
        "３  年次別 都道府県別 罪種別 認知・検挙件数及び検挙人員"
        if valid_title
        else "unrelated table"
    )
    worksheet["C4"] = "刑法犯総数（交通業過を除く）"
    worksheet["B5"] = "年次及び都道府県"
    worksheet["C5"] = "認知件数"
    worksheet["E5"] = "検挙件数"
    worksheet["F5"] = "検挙人員"

    worksheet["B9"] = "2023 令和５年"
    worksheet["C9"] = 90
    worksheet["E9"] = 40
    worksheet["F9"] = 30
    worksheet["B10"] = "2024 令和６年"
    worksheet["C10"] = 100
    worksheet["E10"] = 50
    worksheet["F10"] = 35

    worksheet["B12"] = "北 海 道"
    worksheet["C12"] = 10
    worksheet["E12"] = 6
    worksheet["F12"] = 4
    worksheet["B13"] = "札　　幌"
    worksheet["C13"] = 7
    worksheet["E13"] = 4
    worksheet["F13"] = 3
    worksheet["B14"] = "東　　北"
    worksheet["C14"] = 20
    worksheet["E14"] = 12
    worksheet["F14"] = 8
    worksheet["B15"] = "青    森"
    worksheet["C15"] = 3
    worksheet["E15"] = 2
    worksheet["F15"] = 1
    worksheet["B16"] = "東　　京"
    worksheet["C16"] = 30
    worksheet["E16"] = 14
    worksheet["F16"] = 10
    worksheet["B17"] = "注　fixture"
    workbook.save(path)
    return path


def _population_fixture(path: Path, *, valid_title: bool = True) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "01"
    worksheet["F2"] = "144　年次別" if valid_title else "unrelated table"
    worksheet["L2"] = "都道府県別　人口"
    worksheet["O3"] = "(単位:1,000人）"
    worksheet["C4"] = "都 道 府 県"
    worksheet["N4"] = "令和５年"
    worksheet["O4"] = "令和６年"
    worksheet["N5"] = "2023年"
    worksheet["O5"] = "2024年"
    worksheet["C6"] = "総人口"
    worksheet["N6"] = 123000
    worksheet["O6"] = 122000
    worksheet["C7"] = "北海道"
    worksheet["N7"] = 5100
    worksheet["O7"] = 5000
    worksheet["C8"] = "青森"
    worksheet["N8"] = 1200
    worksheet["O8"] = 1100
    worksheet["C9"] = "東京"
    worksheet["N9"] = 14000
    worksheet["O9"] = 14100
    worksheet["C10"] = "注１　総務省統計局の人口推計及び国勢調査人口（各年１０月１日現在）である。"
    workbook.save(path)
    return path


def _japanese_population_fixture(path: Path, *, valid_title: bool = True) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "第2表"
    worksheet["A1"] = (
        "第２表 都道府県、男女別人口及び人口性比－総人口、日本人人口（2024年10月１日現在）"
        if valid_title
        else "unrelated table"
    )
    worksheet["L4"] = "（単位 千人）"
    worksheet["E6"] = "総人口"
    worksheet["I6"] = "日本人人口"
    worksheet["E9"] = "男女計"
    worksheet["I9"] = "男女計"
    worksheet["A12"] = "全国"
    worksheet["E12"] = 123802
    worksheet["I12"] = 120296
    worksheet["B13"] = "01"
    worksheet["C13"] = "北海道"
    worksheet["E13"] = 5043
    worksheet["I13"] = 4980
    worksheet["B14"] = "13"
    worksheet["C14"] = "東京都"
    worksheet["E14"] = 14178
    worksheet["I14"] = 13463
    worksheet["A15"] = "注 fixture"
    workbook.save(path)
    return path


def _find(records, *, geography: str):
    return next(record for record in records if record.geography == geography)


def test_overall_crime_parser_preserves_total_and_police_geography(tmp_path):
    path = _crime_fixture(tmp_path / "R06_003.xlsx")

    records = parse_npa_overall_prefecture_crime(path, source_id="S15")

    national = _find(records, geography="日本")
    hokkaido = _find(records, geography="北海道")
    sapporo = _find(records, geography="札幌")
    tohoku = _find(records, geography="東北")
    aomori = _find(records, geography="青森県")
    tokyo = _find(records, geography="東京都")
    assert (national.recognized_cases, national.cleared_cases, national.cleared_persons) == (
        100,
        50,
        35,
    )
    assert national.year == 2024
    assert national.population_scope == "all_persons"
    assert national.offense_scope == "criminal_code_excluding_traffic_negligence"
    assert national.geography_type == "national"
    assert hokkaido.geography_type == "prefecture"
    assert sapporo.geography_type == "police_subregion"
    assert sapporo.parent_region == "北海道"
    assert tohoku.geography_type == "police_region"
    assert aomori.parent_region == "東北"
    assert tokyo.geography_type == "prefecture"
    assert tokyo.geography_semantics == "police_reporting_area_unresolved"
    assert tokyo.source_table == "3"


def test_overall_crime_parser_stops_before_official_verification_block(tmp_path):
    path = _crime_fixture(tmp_path / "R02_003.xlsx")
    workbook = load_workbook(path)
    worksheet = workbook["刑法犯総数"]
    worksheet["B17"] = "確認用"
    worksheet["B18"] = "総数"
    worksheet["C18"] = 0
    worksheet["E18"] = 0
    worksheet["F18"] = 0
    workbook.save(path)

    records = parse_npa_overall_prefecture_crime(path, source_id="S15_R02")

    assert len(records) == 6
    assert {record.year for record in records} == {2024}
    assert not any(record.geography == "確認用" for record in records)


def test_overall_crime_parser_rejects_unrelated_workbook(tmp_path):
    path = _crime_fixture(tmp_path / "wrong.xlsx", valid_title=False)

    with pytest.raises(SchemaError, match="Table 3 title"):
        parse_npa_overall_prefecture_crime(path, source_id="S15")


def test_population_parser_converts_published_thousands_without_hiding_rounding(
    tmp_path,
):
    path = _population_fixture(tmp_path / "R06_144.xlsx")

    records = parse_npa_prefecture_population(path, source_id="S16")

    national = _find(records, geography="日本")
    tokyo = _find(records, geography="東京都")
    assert national.year == 2024
    assert national.reference_date == "2024-10-01"
    assert national.population_scope == "total_population"
    assert national.population == 122000000
    assert national.source_value == 122000
    assert national.source_unit == "1000_persons"
    assert national.rounding == "nearest_1000_persons"
    assert national.geography_type == "national"
    assert tokyo.population == 14100000
    assert tokyo.geography_type == "prefecture"
    assert tokyo.geography_semantics == "population_estimate_prefecture"
    assert tokyo.source_table == "144"


def test_population_parser_rejects_unrelated_workbook(tmp_path):
    path = _population_fixture(tmp_path / "wrong.xlsx", valid_title=False)

    with pytest.raises(SchemaError, match="Table 144 title"):
        parse_npa_prefecture_population(path, source_id="S16")


def test_japanese_population_parser_preserves_published_scope_and_rounding(tmp_path):
    path = _japanese_population_fixture(tmp_path / "05k2024-2.xlsx")

    records = parse_statistics_bureau_japanese_population(path, source_id="S17")

    national = _find(records, geography="日本")
    tokyo = _find(records, geography="東京都")
    assert len(records) == 3
    assert national.reference_date == "2024-10-01"
    assert national.population_scope == "japanese_population"
    assert national.population == 120_296_000
    assert national.source_value == 120_296
    assert national.source_unit == "1000_persons"
    assert national.rounding == "nearest_1000_persons"
    assert tokyo.population == 13_463_000
    assert tokyo.geography_type == "prefecture"
    assert tokyo.source_table == "2"


def test_japanese_population_parser_rejects_unrelated_workbook(tmp_path):
    path = _japanese_population_fixture(
        tmp_path / "wrong.xlsx", valid_title=False
    )

    with pytest.raises(SchemaError, match="Statistics Bureau Table 2 title"):
        parse_statistics_bureau_japanese_population(path, source_id="S17")


@pytest.mark.parametrize(
    ("parser", "source_id", "source_table", "fixture_factory", "record_type"),
    [
        (
            "npa-overall-prefecture-crime",
            "S15",
            "3",
            _crime_fixture,
            "overall_prefecture_crime",
        ),
        (
            "npa-prefecture-population",
            "S16",
            "144",
            _population_fixture,
            "prefecture_population",
        ),
        (
            "statistics-bureau-japanese-population",
            "S17",
            "2",
            _japanese_population_fixture,
            "prefecture_population",
        ),
    ],
)
def test_pipeline_dispatches_all_resident_context_parsers(
    parser,
    source_id,
    source_table,
    fixture_factory,
    record_type,
    tmp_path,
):
    source_path = fixture_factory(tmp_path / (source_id + ".xlsx"))
    expected_count = {"S15": 6, "S16": 4, "S17": 3}[source_id]
    profile = {
        "record_type": record_type,
        "expected_record_count": expected_count,
        "expected_years": [2024],
        "allowed_values": {},
        "expected_distinct_counts": {},
        "expected_sums": {},
        "anchors": [],
    }
    metadata = {
        "series_id": "fixture-series",
        "edition_id": source_id,
        "publisher": "Official publisher",
        "dataset": "Fixture dataset",
        "source_table": source_table,
        "parser": parser,
        "expected_format": "xlsx",
        "landing_url": "https://example.test/landing",
        "download_url": "https://example.test/download",
        "period": "2024",
        "license_url": "https://example.test/terms",
        "notes": ["Fixture"],
    }

    result = run_offline_pipeline(
        source_path,
        source_id=source_id,
        source_metadata=metadata,
        quality_profile=profile,
        retrieved_at="2026-09-01T01:00:00+09:00",
        raw_root=tmp_path / "raw",
        processed_root=tmp_path / "processed",
    )

    assert result.reused is False
    assert result.normalized_path.read_text(encoding="utf-8").count("\n") == expected_count
    run_manifest = json.loads(result.run_manifest_path.read_text(encoding="utf-8"))
    assert run_manifest["parser_contract_version"] == (
        2 if source_id == "S17" else 1
    )


@pytest.mark.parametrize(
    ("command", "source_id", "fixture_factory"),
    [
        ("npa-overall-prefecture-crime", "S15", _crime_fixture),
        ("npa-prefecture-population", "S16", _population_fixture),
        (
            "statistics-bureau-japanese-population",
            "S17",
            _japanese_population_fixture,
        ),
    ],
)
def test_ingest_cli_exposes_all_resident_context_parsers(
    command,
    source_id,
    fixture_factory,
    tmp_path,
):
    source_path = fixture_factory(tmp_path / (source_id + ".xlsx"))
    output = tmp_path / (source_id + ".jsonl")
    manifest = tmp_path / (source_id + ".manifest.json")

    result = ingest_main(
        [
            command,
            str(source_path),
            "--source-id",
            source_id,
            "--landing-url",
            "https://example.test/landing",
            "--download-url",
            "https://example.test/download",
            "--retrieved-at",
            "2026-09-01T01:00:00+09:00",
            "--output",
            str(output),
            "--manifest",
            str(manifest),
        ]
    )

    assert result == 0
    assert output.exists()
    assert manifest.exists()
