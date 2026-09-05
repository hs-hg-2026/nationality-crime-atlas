import json
import shutil
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from nationality_crime_atlas.clearance_share_trend import (
    generate_clearance_share_trend,
)
from nationality_crime_atlas.errors import IntegrityError, SchemaError
from nationality_crime_atlas.provenance import sha256_file


def _write_visiting_foreign_fixture(path: Path, *, cases_2024: int = 40) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "01 "
    worksheet.cell(2, 4, "131 罪種別 国籍別 来日外国人による犯罪")
    worksheet.cell(4, 6, "刑法犯")
    worksheet.cell(5, 6, "計")
    worksheet.cell(7, 6, "件数")
    worksheet.cell(7, 7, "人員")
    worksheet.cell(8, 3, "2023年")
    worksheet.cell(8, 6, 30)
    worksheet.cell(8, 7, 20)
    worksheet.cell(9, 3, "2024年")
    worksheet.cell(9, 6, cases_2024)
    worksheet.cell(9, 7, 25)
    workbook.save(path)


def _write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _build_inputs(
    tmp_path: Path,
    nationality_table130_file: Path,
    all_person_offense_file: Path,
    *,
    visiting_2024_cases: int = 40,
):
    raw_root = tmp_path / "raw"
    catalog_rows = []
    source_files = {
        "S08": (nationality_table130_file, "130"),
        "S09": (tmp_path / "visiting.xlsx", "131"),
        "S15": (all_person_offense_file, "3"),
    }
    _write_visiting_foreign_fixture(
        source_files["S09"][0], cases_2024=visiting_2024_cases
    )

    workbook = load_workbook(all_person_offense_file)
    workbook["刑法犯総数"].cell(
        2,
        2,
        "3 年次別 都道府県別 罪種別 認知・検挙件数及び検挙人員",
    )
    workbook.save(all_person_offense_file)

    pins = {}
    for source_id, (source_path, source_table) in source_files.items():
        relative = Path("fixture") / source_id / source_path.name
        destination = raw_root / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(source_path, destination)
        pins[source_id] = sha256_file(destination)
        catalog_rows.append(
            {
                "source_id": source_id,
                "series_id": "fixture-%s" % source_id.lower(),
                "publisher": "Fixture official publisher",
                "dataset": "Fixture %s" % source_id,
                "source_table": source_table,
                "source_period": "2013--2024 annual fixture",
                "landing_url": "https://example.test/%s" % source_id.lower(),
                "download_url": "https://example.test/%s.xlsx" % source_id.lower(),
                "retrieved_at": "2026-09-05T18:00:00+09:00",
                "revision": "fixture",
                "verification_level": "binary_and_primary",
                "raw_relpath": relative.as_posix(),
                "sha256": pins[source_id],
                "processing_status": "validated",
            }
        )

    catalog = tmp_path / "catalog.jsonl"
    catalog.write_text(
        "".join(json.dumps(row, sort_keys=True) + "\n" for row in catalog_rows),
        encoding="utf-8",
    )
    contract = tmp_path / "contract.json"
    _write_json(
        contract,
        {
            "schema_version": 1,
            "artifact_pins": pins,
            "trend": {
                "trend_id": "national_criminal_code_clearance_foreign_share",
                "label_ja": "全国の刑法犯検挙に占める外国人区分の割合",
                "label_en": "Foreign-scope share of national criminal-code clearances",
                "years": [2023, 2024],
                "all_person_source_id": "S15",
                "foreign_sources": [
                    {
                        "source_id": "S08",
                        "source_table": "130",
                        "foreign_scope": "all_foreign",
                        "label_ja": "外国人全体",
                    },
                    {
                        "source_id": "S09",
                        "source_table": "131",
                        "foreign_scope": "visiting_foreign",
                        "label_ja": "来日外国人",
                    },
                ],
                "metrics": ["cleared_cases", "cleared_persons"],
                "display_multiplier": 100,
                "display_unit_label_ja": "%",
                "interpretation_policy": "share_of_clearances_not_population_risk",
                "ui_caveat": "検挙全体に占める構成比であり、人口当たりの犯罪率ではない。",
            },
        },
    )
    return catalog, raw_root, contract


def test_clearance_share_trend_keeps_direct_scopes_and_their_arithmetic_residual(
    tmp_path,
    nationality_table130_file,
    all_person_offense_file,
):
    catalog, raw_root, contract = _build_inputs(
        tmp_path,
        nationality_table130_file,
        all_person_offense_file,
    )

    report = generate_clearance_share_trend(
        catalog_path=catalog,
        raw_root=raw_root,
        contract_path=contract,
        output_root=tmp_path / "trend",
        generated_at="2026-09-05T18:10:00+09:00",
    )

    rows = [json.loads(line) for line in report.jsonl_path.read_text().splitlines()]
    assert report.record_count == 12
    assert report.year_count == 2
    assert report.latest_path.is_file()
    assert {
        (row["year"], row["foreign_scope"], row["metric"])
        for row in rows
    } == {
        (year, scope, metric)
        for year in (2023, 2024)
        for scope in (
            "all_foreign",
            "visiting_foreign",
            "all_foreign_minus_visiting_foreign",
        )
        for metric in ("cleared_cases", "cleared_persons")
    }
    all_foreign_cases = next(
        row
        for row in rows
        if row["year"] == 2024
        and row["foreign_scope"] == "all_foreign"
        and row["metric"] == "cleared_cases"
    )
    assert all_foreign_cases["numerator_value"] == 60
    assert all_foreign_cases["denominator_value"] == 600
    assert all_foreign_cases["display_value"] == pytest.approx(10.0)
    visiting_persons = next(
        row
        for row in rows
        if row["year"] == 2024
        and row["foreign_scope"] == "visiting_foreign"
        and row["metric"] == "cleared_persons"
    )
    assert visiting_persons["numerator_value"] == 25
    assert visiting_persons["denominator_value"] == 300
    assert visiting_persons["display_value"] == pytest.approx(8.333333333333332)
    assert "visiting_foreign_includes_nonresidents" in visiting_persons["mismatch_flags"]

    residual_cases = next(
        row
        for row in rows
        if row["year"] == 2024
        and row["foreign_scope"] == "all_foreign_minus_visiting_foreign"
        and row["metric"] == "cleared_cases"
    )
    assert residual_cases["foreign_scope_label_ja"] == (
        "外国人全体−来日外国人（差分）"
    )
    assert residual_cases["numerator_value"] == 20
    assert residual_cases["denominator_value"] == 600
    assert residual_cases["display_value"] == pytest.approx(20 / 600 * 100)
    assert residual_cases["numerator_source_ids"] == ["S08", "S09"]
    assert residual_cases["derivation_method"] == (
        "arithmetic_residual_all_foreign_minus_visiting_foreign"
    )
    assert "residual_not_equivalent_to_usual_residents" in residual_cases[
        "mismatch_flags"
    ]


def test_clearance_share_trend_refuses_a_negative_foreign_scope_residual(
    tmp_path,
    nationality_table130_file,
    all_person_offense_file,
):
    catalog, raw_root, contract = _build_inputs(
        tmp_path,
        nationality_table130_file,
        all_person_offense_file,
        visiting_2024_cases=61,
    )

    with pytest.raises(
        SchemaError,
        match="Visiting-foreign clearances exceed all-foreign clearances",
    ):
        generate_clearance_share_trend(
            catalog_path=catalog,
            raw_root=raw_root,
            contract_path=contract,
            output_root=tmp_path / "trend",
            generated_at="2026-09-05T18:10:00+09:00",
        )


def test_clearance_share_trend_stops_on_unreviewed_artifact_hash(
    tmp_path,
    nationality_table130_file,
    all_person_offense_file,
):
    catalog, raw_root, contract = _build_inputs(
        tmp_path,
        nationality_table130_file,
        all_person_offense_file,
    )
    data = json.loads(contract.read_text(encoding="utf-8"))
    data["artifact_pins"]["S09"] = "0" * 64
    _write_json(contract, data)

    with pytest.raises(IntegrityError, match="artifact pin"):
        generate_clearance_share_trend(
            catalog_path=catalog,
            raw_root=raw_root,
            contract_path=contract,
            output_root=tmp_path / "trend",
            generated_at="2026-09-05T18:10:00+09:00",
        )
