import json
import shutil
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from nationality_crime_atlas.clearance_population_trend import (
    generate_clearance_population_trend,
)
from nationality_crime_atlas.errors import IntegrityError, SchemaError
from nationality_crime_atlas.provenance import sha256_file


def _write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _write_japanese_population_fixture(
    path: Path,
    *,
    year: int,
    population_thousands: int,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "第2表"
    worksheet["A1"] = (
        "第２表 都道府県、男女別人口及び人口性比－総人口、日本人人口"
        f"（{year}年10月１日現在）"
    )
    worksheet["L4"] = "（単位 千人）"
    worksheet["E6"] = "総人口"
    worksheet["I6"] = "日本人人口"
    worksheet["I9"] = "男女計"
    worksheet["A12"] = "全国"
    worksheet["I12"] = population_thousands
    workbook.save(path)


def _write_foreign_population_fixture(
    path: Path,
    *,
    year: int,
    population: int,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"{str(year)[-2:]}-12-01m"
    worksheet.append(("統計名：", "在留外国人統計"))
    worksheet.append(("統計表番号：", "第1表"))
    worksheet.append(("表題：", "国籍・地域別 在留資格別 在留外国人"))
    worksheet.append(("時点", "州", "国籍・地域", "在留資格", "在留外国人数"))
    worksheet.append((f"{year}年12月末", "総数", "総数", "総数", population))
    worksheet.append((f"{year}年12月末", "アジア", "ベトナム", "総数", population))
    workbook.save(path)


def _build_inputs(
    tmp_path: Path,
    nationality_table130_file: Path,
    all_person_offense_file: Path,
    *,
    japanese_2024_year: int = 2024,
):
    workbook = load_workbook(all_person_offense_file)
    workbook["刑法犯総数"].cell(
        2,
        2,
        "3 年次別 都道府県別 罪種別 認知・検挙件数及び検挙人員",
    )
    workbook.save(all_person_offense_file)

    japanese_2023 = tmp_path / "japanese-2023.xlsx"
    japanese_2024 = tmp_path / "japanese-2024.xlsx"
    foreign_2024 = tmp_path / "foreign-2024.xlsx"
    _write_japanese_population_fixture(
        japanese_2023,
        year=2023,
        population_thousands=121_000,
    )
    _write_japanese_population_fixture(
        japanese_2024,
        year=japanese_2024_year,
        population_thousands=120_000,
    )
    _write_foreign_population_fixture(
        foreign_2024,
        year=2024,
        population=100_000,
    )

    source_files = {
        "S08": (nationality_table130_file, "130"),
        "S15": (all_person_offense_file, "3"),
        "S17_2023": (japanese_2023, "2"),
        "S17": (japanese_2024, "2"),
        "S19_2024": (foreign_2024, "1"),
    }
    raw_root = tmp_path / "raw"
    pins = {}
    catalog_rows = []
    for source_id, (source_path, source_table) in source_files.items():
        relative = Path("fixture") / source_id / source_path.name
        destination = raw_root / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(source_path, destination)
        pins[source_id] = sha256_file(destination)
        catalog_rows.append(
            {
                "source_id": source_id,
                "series_id": "fixture-%s" % source_id.lower(),
                "publisher": "Fixture official publisher",
                "dataset": "Fixture %s" % source_id,
                "source_table": source_table,
                "source_period": "fixture period",
                "landing_url": "https://example.test/%s" % source_id.lower(),
                "download_url": "https://example.test/%s.xlsx" % source_id.lower(),
                "retrieved_at": "2026-09-06T10:00:00+09:00",
                "revision": "fixture",
                "verification_level": "binary_and_primary",
                "raw_relpath": relative.as_posix(),
                "sha256": pins[source_id],
                "processing_status": "validated",
            }
        )

    catalog = tmp_path / "catalog.jsonl"
    catalog.write_text(
        "".join(json.dumps(row, sort_keys=True) + "\n" for row in catalog_rows),
        encoding="utf-8",
    )
    contract = tmp_path / "contract.json"
    _write_json(
        contract,
        {
            "schema_version": 1,
            "artifact_pins": pins,
            "trend": {
                "trend_id": "national_clearance_population_reference_ratio",
                "label_ja": "人口1,000人当たりの刑法犯検挙参考比率",
                "label_en": "Criminal-code clearances per 1,000 reference population",
                "years": [2023, 2024],
                "all_person_source_id": "S15",
                "all_foreign_source_id": "S08",
                "japanese_population_sources": {
                    "2023": "S17_2023",
                    "2024": "S17",
                },
                "resident_foreign_population_sources": {
                    "2024": "S19_2024",
                },
                "metrics": ["cleared_cases", "cleared_persons"],
                "display_multiplier": 1000,
                "display_unit_label_ja": "人口1,000人当たり",
                "interpretation_policy": "public_data_reference_ratio_not_probability",
                "ui_caveat": "分子と分母の対象範囲と基準日が一致しない参考比率である。",
            },
        },
    )
    return catalog, raw_root, contract


def test_clearance_population_trend_keeps_counts_populations_rates_and_refusals(
    tmp_path,
    nationality_table130_file,
    all_person_offense_file,
):
    catalog, raw_root, contract = _build_inputs(
        tmp_path,
        nationality_table130_file,
        all_person_offense_file,
    )

    report = generate_clearance_population_trend(
        catalog_path=catalog,
        raw_root=raw_root,
        contract_path=contract,
        output_root=tmp_path / "trend",
        generated_at="2026-09-06T10:10:00+09:00",
    )

    rows = [json.loads(line) for line in report.jsonl_path.read_text().splitlines()]
    assert report.record_count == 8
    assert report.calculated_count == 6
    assert report.refused_count == 2
    assert {
        (row["year"], row["population_group"], row["metric"])
        for row in rows
    } == {
        (year, group, metric)
        for year in (2023, 2024)
        for group in ("japanese_etc_residual", "all_foreign")
        for metric in ("cleared_cases", "cleared_persons")
    }

    japanese_cases = next(
        row
        for row in rows
        if row["year"] == 2024
        and row["population_group"] == "japanese_etc_residual"
        and row["metric"] == "cleared_cases"
    )
    assert japanese_cases["numerator_value"] == 540
    assert japanese_cases["denominator_value"] == 120_000_000
    assert japanese_cases["display_value"] == pytest.approx(0.0045)
    assert japanese_cases["numerator_source_ids"] == ["S15", "S08"]
    assert japanese_cases["denominator_source_id"] == "S17"
    assert japanese_cases["population_reference_date"] == "2024-10-01"
    assert japanese_cases["derivation_formula"] == (
        "(S15.cleared_cases - S08.cleared_cases) / S17.population * 1000"
    )
    assert "japanese_numerator_is_arithmetic_residual" in japanese_cases[
        "mismatch_flags"
    ]
    assert len(japanese_cases["source_components"]) == 3

    foreign_persons = next(
        row
        for row in rows
        if row["year"] == 2024
        and row["population_group"] == "all_foreign"
        and row["metric"] == "cleared_persons"
    )
    assert foreign_persons["numerator_value"] == 40
    assert foreign_persons["denominator_value"] == 100_000
    assert foreign_persons["display_value"] == pytest.approx(0.4)
    assert foreign_persons["numerator_source_ids"] == ["S08"]
    assert foreign_persons["denominator_source_id"] == "S19_2024"
    assert foreign_persons["population_reference_date"] == "2024-12-31"
    assert "all_foreign_numerator_vs_resident_foreigner_denominator" in (
        foreign_persons["mismatch_flags"]
    )

    foreign_2023 = next(
        row
        for row in rows
        if row["year"] == 2023
        and row["population_group"] == "all_foreign"
        and row["metric"] == "cleared_cases"
    )
    assert foreign_2023["calculation_status"] == "refused"
    assert foreign_2023["numerator_value"] == 55
    assert foreign_2023["denominator_value"] is None
    assert foreign_2023["display_value"] is None
    assert foreign_2023["denominator_source_id"] is None
    assert foreign_2023["refusal_reason"] == (
        "resident_foreigner_population_source_not_registered_for_year"
    )
    assert len(foreign_2023["source_components"]) == 1


def test_clearance_population_trend_rejects_negative_japanese_residual(
    tmp_path,
    nationality_table130_file,
    all_person_offense_file,
):
    workbook = load_workbook(nationality_table130_file)
    workbook["01"].cell(9, 7, 601)
    workbook.save(nationality_table130_file)
    catalog, raw_root, contract = _build_inputs(
        tmp_path,
        nationality_table130_file,
        all_person_offense_file,
    )

    with pytest.raises(SchemaError, match="exceed all-person clearances"):
        generate_clearance_population_trend(
            catalog_path=catalog,
            raw_root=raw_root,
            contract_path=contract,
            output_root=tmp_path / "trend",
            generated_at="2026-09-06T10:10:00+09:00",
        )


def test_clearance_population_trend_rejects_population_year_mismatch(
    tmp_path,
    nationality_table130_file,
    all_person_offense_file,
):
    catalog, raw_root, contract = _build_inputs(
        tmp_path,
        nationality_table130_file,
        all_person_offense_file,
        japanese_2024_year=2023,
    )

    with pytest.raises(SchemaError, match="population year"):
        generate_clearance_population_trend(
            catalog_path=catalog,
            raw_root=raw_root,
            contract_path=contract,
            output_root=tmp_path / "trend",
            generated_at="2026-09-06T10:10:00+09:00",
        )


def test_clearance_population_trend_stops_on_unreviewed_artifact_hash(
    tmp_path,
    nationality_table130_file,
    all_person_offense_file,
):
    catalog, raw_root, contract = _build_inputs(
        tmp_path,
        nationality_table130_file,
        all_person_offense_file,
    )
    data = json.loads(contract.read_text(encoding="utf-8"))
    data["artifact_pins"]["S19_2024"] = "0" * 64
    _write_json(contract, data)

    with pytest.raises(IntegrityError, match="artifact pin"):
        generate_clearance_population_trend(
            catalog_path=catalog,
            raw_root=raw_root,
            contract_path=contract,
            output_root=tmp_path / "trend",
            generated_at="2026-09-06T10:10:00+09:00",
        )
