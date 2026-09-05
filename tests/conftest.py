from pathlib import Path

import pytest
import xlwt
from openpyxl import Workbook


POPULATION_HEADERS = [
    "国籍・地域",
    "在留資格",
    "性別",
    "年齢（５歳階級）",
    "年齢",
    "都道府県",
    "在留外国人数",
]


@pytest.fixture
def population_t1_file(tmp_path: Path) -> Path:
    path = tmp_path / "population_t1.xlsx"
    workbook = Workbook()
    pivot = workbook.active
    pivot.title = "PVT"
    pivot.append(["国籍・地域", "(すべて)"])

    rows = workbook.create_sheet("令和7年12月末")
    rows.sheet_state = "hidden"
    rows.append(POPULATION_HEADERS)
    rows.append(
        [
            "01_011：韓国",
            "35：永住者",
            "2:女",
            "17_80歳以上",
            "80歳以上",
            "21：岐阜県",
            15,
        ]
    )
    rows.append(
        [
            "01_022：台湾",
            "36：日本人の配偶者等",
            "1:男",
            "16_75～79歳",
            "75歳",
            "13：東京都",
            "-",
        ]
    )
    workbook.save(path)
    return path


@pytest.fixture
def malformed_population_file(tmp_path: Path) -> Path:
    path = tmp_path / "malformed_population.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "令和7年12月末"
    worksheet.append(["国籍・地域", "都道府県", "在留外国人数"])
    worksheet.append(["韓国", "岐阜県", 15])
    workbook.save(path)
    return path


@pytest.fixture
def nationality_population_totals_file(tmp_path: Path) -> Path:
    path = tmp_path / "24-12-01-1.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "24-12-01m"
    worksheet.append(("統計名：", "在留外国人統計"))
    worksheet.append(("統計表番号：", "第1表"))
    worksheet.append(("表題：", "国籍・地域別 在留資格別 在留外国人"))
    worksheet.append(("時点", "州", "国籍・地域", "在留資格", "在留外国人数"))
    worksheet.append(("令和6年12月末", "総数", "総数", "総数", 100))
    worksheet.append(("令和6年12月末", "アジア", "総数", "総数", 80))
    worksheet.append(("令和6年12月末", "アジア", "ベトナム", "総数", 50))
    worksheet.append(("令和6年12月末", "アジア", "中国", "総数", 30))
    worksheet.append(("令和6年12月末", "北アメリカ", "総数", "総数", 20))
    worksheet.append(("令和6年12月末", "北アメリカ", "米国", "総数", 20))
    worksheet.append(("令和6年12月末", "無国籍", "総数", "総数", 0))
    worksheet.append(("令和6年12月末", "アジア", "うち中国〔香港〕", "総数", 5))
    workbook.save(path)
    return path


def _write_nationality_fixture(path: Path, table_id: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "01"

    if table_id == "130":
        worksheet.cell(2, 4, "130 罪種別 国籍別 外国人による犯罪")
        label_column = 4
        metric_column = 5
        worksheet.cell(7, metric_column, "件数")
        worksheet.cell(7, metric_column + 1, "人員")
    else:
        worksheet.cell(2, 4, "131 罪種別 国籍別 来日外国人による犯罪")
        label_column = 3
        metric_column = 4
        worksheet.cell(7, metric_column, "件数")
        worksheet.cell(7, metric_column + 1, "人員")

    criminal_code_metric_column = metric_column + 2
    worksheet.cell(4, criminal_code_metric_column, "刑法犯")
    worksheet.cell(5, criminal_code_metric_column, "計")
    worksheet.cell(7, criminal_code_metric_column, "件数")
    worksheet.cell(7, criminal_code_metric_column + 1, "人員")

    worksheet.cell(8, label_column, "2023年")
    worksheet.cell(8, metric_column, 90)
    worksheet.cell(8, metric_column + 1, 60)
    worksheet.cell(8, criminal_code_metric_column, 55)
    worksheet.cell(8, criminal_code_metric_column + 1, 35)
    worksheet.cell(9, label_column, "2024年")
    worksheet.cell(9, metric_column, 100)
    worksheet.cell(9, metric_column + 1, 70)
    worksheet.cell(9, criminal_code_metric_column, 60)
    worksheet.cell(9, criminal_code_metric_column + 1, 40)

    worksheet.cell(12, 2, "アジア州の国")
    worksheet.cell(12, metric_column, 80)
    worksheet.cell(12, metric_column + 1, 50)
    worksheet.cell(12, criminal_code_metric_column, 48)
    worksheet.cell(12, criminal_code_metric_column + 1, 30)
    worksheet.cell(13, 3, "中国")
    worksheet.cell(13, metric_column, 30)
    worksheet.cell(13, metric_column + 1, 20)
    worksheet.cell(13, criminal_code_metric_column, 18)
    worksheet.cell(13, criminal_code_metric_column + 1, 12)
    worksheet.cell(14, 3, "ベトナム")
    worksheet.cell(14, metric_column, 40)
    worksheet.cell(14, metric_column + 1, 25)
    worksheet.cell(14, criminal_code_metric_column, 24)
    worksheet.cell(14, criminal_code_metric_column + 1, 15)

    if table_id == "130":
        worksheet.cell(15, 2, "南北アメリカ州の国")
        worksheet.cell(15, metric_column, 20)
        worksheet.cell(15, metric_column + 1, 15)
        worksheet.cell(15, criminal_code_metric_column, 12)
        worksheet.cell(15, criminal_code_metric_column + 1, 9)
        worksheet.cell(16, 2, "アメリカ")
        worksheet.cell(16, 4, "軍人")
        worksheet.cell(16, metric_column, 3)
        worksheet.cell(16, metric_column + 1, 2)
        worksheet.cell(16, criminal_code_metric_column, 2)
        worksheet.cell(16, criminal_code_metric_column + 1, 1)
        worksheet.cell(17, 4, "その他")
        worksheet.cell(17, metric_column, 7)
        worksheet.cell(17, metric_column + 1, 6)
        worksheet.cell(17, criminal_code_metric_column, 4)
        worksheet.cell(17, criminal_code_metric_column + 1, 3)
        note_row = 18
    else:
        note_row = 15

    worksheet.cell(note_row, 2, "注 「中国」には、台湾、香港等を含む。")
    workbook.save(path)


@pytest.fixture(params=["130", "131"])
def nationality_crime_file(tmp_path: Path, request):
    table_id = request.param
    path = tmp_path / ("table_%s.xlsx" % table_id)
    _write_nationality_fixture(path, table_id)
    return table_id, path


@pytest.fixture
def nationality_table130_file(tmp_path: Path) -> Path:
    path = tmp_path / "table_130.xlsx"
    _write_nationality_fixture(path, "130")
    return path


@pytest.fixture
def malformed_nationality_file(tmp_path: Path) -> Path:
    path = tmp_path / "malformed_nationality.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "01"
    worksheet.cell(2, 4, "130 罪種別 国籍別 外国人による犯罪")
    worksheet.cell(4, 2, "no metric headers")
    workbook.save(path)
    return path


OFFENSE_GROUP_COLUMNS = {
    "all_offenses": ("01 ", 5),
    "criminal_code": ("01 ", 7),
    "heinous": ("01 ", 9),
    "assaultive": ("01 ", 20),
    "theft": ("02", 9),
    "intellectual": ("02", 13),
    "morals": ("03 ", 7),
    "other_criminal_code": ("03 ", 16),
    "special_law": ("04 ", 5),
}

OFFENSE_GROUP_LABELS = {
    "all_offenses": "総数",
    "criminal_code": "刑法犯",
    "heinous": "凶悪犯",
    "assaultive": "粗暴犯",
    "theft": "窃盗犯",
    "intellectual": "知能犯",
    "morals": "風俗犯",
    "other_criminal_code": "その他の刑法犯",
    "special_law": "特別法犯",
}


def _offense_values(scale: int):
    persons = {
        "heinous": scale,
        "assaultive": scale * 2,
        "theft": scale * 3,
        "intellectual": scale * 2,
        "morals": scale,
        "other_criminal_code": scale,
    }
    cases = {key: value * 2 for key, value in persons.items()}
    persons["criminal_code"] = sum(persons.values())
    cases["criminal_code"] = sum(cases.values())
    persons["special_law"] = scale * 2
    cases["special_law"] = scale * 3
    persons["all_offenses"] = persons["criminal_code"] + persons["special_law"]
    cases["all_offenses"] = cases["criminal_code"] + cases["special_law"]
    return cases, persons


def _write_offense_headers(worksheet, table_id: str, sheet_name: str) -> None:
    worksheet.cell(2, 4, f"{table_id} 罪種別 国籍別 外国人による犯罪")
    header_row = 8 if sheet_name == "03 " else 7
    for offense_id, (target_sheet, cases_column) in OFFENSE_GROUP_COLUMNS.items():
        if target_sheet != sheet_name:
            continue
        worksheet.cell(header_row, cases_column, "件数")
        worksheet.cell(header_row, cases_column + 1, "人員")
        worksheet.cell(4, cases_column, OFFENSE_GROUP_LABELS[offense_id])


def _write_offense_entity_rows(worksheet, sheet_name: str) -> None:
    header_row = 8 if sheet_name == "03 " else 7
    year_row = header_row + 1
    latest_year_row = header_row + 2
    first_entity_row = header_row + 5
    entity_specs = [
        ([(2, "アジア州の国")], 12),
        ([(3, "中国")], 5),
        ([(3, "ベトナム")], 4),
        ([(2, "南北アメリカ州の国")], 4),
        ([(2, "アメリカ"), (4, "軍人")], 1),
        ([(4, "その他")], 3),
        ([(2, "無国籍")], 1),
        ([(2, "国籍不明")], 1),
    ]
    worksheet.cell(year_row, 4, "2023年")
    worksheet.cell(latest_year_row, 4, "2024年")
    annual_cases, annual_persons = _offense_values(18)
    for offense_id, (target_sheet, cases_column) in OFFENSE_GROUP_COLUMNS.items():
        if target_sheet != sheet_name:
            continue
        worksheet.cell(year_row, cases_column, annual_cases[offense_id] - 1)
        worksheet.cell(year_row, cases_column + 1, annual_persons[offense_id] - 1)
        worksheet.cell(latest_year_row, cases_column, annual_cases[offense_id])
        worksheet.cell(latest_year_row, cases_column + 1, annual_persons[offense_id])

    for offset, (labels, scale) in enumerate(entity_specs):
        source_row = first_entity_row + offset
        for label_column, label in labels:
            worksheet.cell(source_row, label_column, label)
        cases, persons = _offense_values(scale)
        for offense_id, (target_sheet, cases_column) in OFFENSE_GROUP_COLUMNS.items():
            if target_sheet != sheet_name:
                continue
            worksheet.cell(source_row, cases_column, cases[offense_id])
            worksheet.cell(source_row, cases_column + 1, persons[offense_id])
    worksheet.cell(first_entity_row + len(entity_specs), 2, "注 fixture")


@pytest.fixture
def nationality_offense_file(tmp_path: Path) -> Path:
    path = tmp_path / "nationality_offenses.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "01 "
    for sheet_name in ("02", "03 ", "04 "):
        workbook.create_sheet(sheet_name)
    for worksheet in workbook.worksheets:
        _write_offense_headers(worksheet, "130", worksheet.title)
        _write_offense_entity_rows(worksheet, worksheet.title)
    workbook.save(path)
    return path


@pytest.fixture
def all_person_offense_file(tmp_path: Path) -> Path:
    path = tmp_path / "all_person_offenses.xlsx"
    workbook = Workbook()
    category_specs = [
        ("刑法犯総数", "刑法犯総数（交通業過を除く）", 300),
        ("A", "A 凶悪犯", 30),
        ("B", "B 粗暴犯", 60),
        ("C", "C 窃盗犯", 90),
        ("D", "D 知能犯", 60),
        ("E", "E 風俗犯", 30),
        ("F", "F その他の刑法犯", 30),
    ]
    for index, (sheet_name, title, persons) in enumerate(category_specs):
        worksheet = workbook.active if index == 0 else workbook.create_sheet()
        worksheet.title = sheet_name
        worksheet.cell(2, 2, "3 年次別 都道府県別 罪種別")
        worksheet.cell(4, 3, title)
        worksheet.cell(5, 2, "年次及び都道府県")
        worksheet.cell(5, 3, "認知件数")
        worksheet.cell(5, 5, "検挙件数")
        worksheet.cell(5, 6, "検挙人員")
        worksheet.cell(9, 2, "2023 令和5年")
        worksheet.cell(9, 3, persons * 3 - 1)
        worksheet.cell(9, 5, persons * 2 - 1)
        worksheet.cell(9, 6, persons - 1)
        worksheet.cell(10, 2, "2024 令和6年")
        worksheet.cell(10, 3, persons * 3)
        worksheet.cell(10, 5, persons * 2)
        worksheet.cell(10, 6, persons)
    workbook.save(path)
    return path


def _write_row(sheet, row_index, values):
    for column_index, value in enumerate(values):
        sheet.write(row_index, column_index, value)


@pytest.fixture
def prefecture_table13_file(tmp_path: Path) -> Path:
    path = tmp_path / "table13.xls"
    workbook = xlwt.Workbook()

    combined = workbook.add_sheet("第１３表＿刑法犯")
    _write_row(combined, 0, ["第１３表", "", "来日外国人による 刑法犯・特別法犯"])
    _write_row(combined, 3, ["", "", "検挙件数", "検挙件数", "", "", "検挙人員", "検挙人員"])
    _write_row(combined, 4, ["", "", "刑法犯\n特別法犯", "刑法犯\n特別法犯", "", "", "刑法犯\n特別法犯", "刑法犯\n特別法犯", "", "", "刑法犯", "刑法犯", "", "", "刑法犯", "刑法犯"])
    _write_row(combined, 5, ["", "", "2025年\n1～12月", "2024年\n1～12月", "", "", "2025年\n1～12月", "2024年\n1～12月", "", "", "2025年\n1～12月", "2024年\n1～12月", "", "", "2025年\n1～12月", "2024年\n1～12月"])
    _write_row(combined, 6, ["総数", "", 100, 90, 0, 0, 70, 65, 0, 0, 60, 55, 0, 0, 40, 35])
    _write_row(combined, 7, ["北海道", "計", 10, 9, 0, 0, 8, 7, 0, 0, 6, 5, 0, 0, 4, 3])
    _write_row(combined, 8, ["北海道", "札幌方面", 7, 6, 0, 0, 5, 4, 0, 0, 4, 3, 0, 0, 3, 2])
    _write_row(combined, 9, ["東北", "計", 20, 18, 0, 0, 12, 11, 0, 0, 10, 9, 0, 0, 6, 5])
    _write_row(combined, 10, ["東北", "青森県", 4, 3, 0, 0, 3, 2, 0, 0, 2, 1, 0, 0, 1, 1])
    _write_row(combined, 11, ["東京都", "", 30, 28, 0, 0, 20, 19, 0, 0, 16, 15, 0, 0, 10, 9])

    special = workbook.add_sheet("第１３表＿特別法犯")
    _write_row(special, 0, ["第１３表", "", "来日外国人による 刑法犯・特別法犯"])
    _write_row(special, 3, ["", "", "検挙件数", "検挙件数", "", "", "検挙人員", "検挙人員"])
    _write_row(special, 4, ["", "", "特別法犯", "特別法犯", "", "", "特別法犯", "特別法犯"])
    _write_row(special, 5, ["", "", "2025年\n1～12月", "2024年\n1～12月", "", "", "2025年\n1～12月", "2024年\n1～12月"])
    _write_row(special, 6, ["総数", "", 40, 35, 0, 0, 30, 30])
    _write_row(special, 7, ["北海道", "計", 4, 4, 0, 0, 4, 4])
    _write_row(special, 8, ["北海道", "札幌方面", 3, 3, 0, 0, 2, 2])
    _write_row(special, 9, ["東北", "計", 10, 9, 0, 0, 6, 6])
    _write_row(special, 10, ["東北", "青森県", 2, 2, 0, 0, 2, 1])
    _write_row(special, 11, ["東京都", "", 14, 13, 0, 0, 10, 10])

    workbook.save(str(path))
    return path


@pytest.fixture
def malformed_prefecture_file(tmp_path: Path) -> Path:
    path = tmp_path / "malformed_table13.xls"
    workbook = xlwt.Workbook()
    workbook.add_sheet("unrelated")
    workbook.save(str(path))
    return path
