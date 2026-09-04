import json
import shutil
from pathlib import Path

import pytest

from nationality_crime_atlas.errors import IntegrityError
from nationality_crime_atlas.offense_composition import (
    generate_offense_composition_report,
    jensen_shannon_distance,
)
from nationality_crime_atlas.provenance import sha256_file


CATEGORIES = [
    ("heinous", "凶悪犯", "#b5443c"),
    ("assaultive", "粗暴犯", "#d17c32"),
    ("theft", "窃盗犯", "#c7a83d"),
    ("intellectual", "知能犯", "#327a92"),
    ("morals", "風俗犯", "#775d9b"),
    ("other_criminal_code", "その他の刑法犯", "#778187"),
]


def _write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _source_row(source_id, raw_relpath, processed_relpath, sha256, table):
    return {
        "artifact_catalog_schema_version": 1,
        "source_id": source_id,
        "series_id": f"series-{source_id.lower()}",
        "publisher": "National Police Agency of Japan",
        "dataset": f"NPA {source_id} fixture",
        "source_table": table,
        "source_period": "2024 annual",
        "coverage_periods": ["2024"],
        "landing_url": "https://example.test/npa",
        "download_url": f"https://example.test/npa/{source_id}.xlsx",
        "retrieved_at": "2026-09-03T20:00:00+09:00",
        "published_at": None,
        "revision": "fixture",
        "verification_level": "binary_and_primary",
        "local_filename": Path(raw_relpath).name,
        "raw_relpath": str(raw_relpath),
        "sha256": sha256,
        "byte_size": 1,
        "file_format": "xlsx",
        "acquisition_mode": "test_fixture",
        "final_url": f"https://example.test/npa/{source_id}.xlsx",
        "http_status": 200,
        "processing_status": "validated",
        "processed_relpath": str(processed_relpath),
        "record_count": 1,
        "quality_passed": True,
    }


def _build_inputs(tmp_path, nationality_offense_file, all_person_offense_file):
    raw_root = tmp_path / "raw"
    processed_root = tmp_path / "processed"
    catalog_path = processed_root / "_catalog" / "artifacts.jsonl"
    source_inputs = {
        "S08": (nationality_offense_file, "series-s08/S08/run/R06_130.xlsx", "130"),
        "S15": (all_person_offense_file, "series-s15/S15/run/R06_003.xlsx", "3"),
    }
    catalog_rows = []
    artifact_pins = {}
    processed_pins = {}
    for source_id, (source, raw_relpath, table) in source_inputs.items():
        raw_path = raw_root / raw_relpath
        raw_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, raw_path)
        artifact_hash = sha256_file(raw_path)
        processed_relpath = Path(raw_relpath).parent
        processed_dir = processed_root / processed_relpath
        normalized_path = processed_dir / "normalized.jsonl"
        normalized_path.parent.mkdir(parents=True, exist_ok=True)
        normalized_path.write_text('{"fixture":true}\n', encoding="utf-8")
        normalized_hash = sha256_file(normalized_path)
        _write_json(
            processed_dir / "run.json",
            {
                "source_id": source_id,
                "raw_artifact_sha256": artifact_hash,
                "normalized_sha256": normalized_hash,
                "quality_passed": True,
            },
        )
        catalog_rows.append(
            _source_row(
                source_id,
                raw_relpath,
                processed_relpath,
                artifact_hash,
                table,
            )
        )
        artifact_pins[source_id] = artifact_hash
        processed_pins[source_id] = normalized_hash

    catalog_path.parent.mkdir(parents=True, exist_ok=True)
    catalog_path.write_text(
        "".join(
            json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n"
            for row in catalog_rows
        ),
        encoding="utf-8",
    )
    contract_path = tmp_path / "offense_contract.json"
    _write_json(
        contract_path,
        {
            "schema_version": 1,
            "artifact_pins": artifact_pins,
            "processed_input_pins": processed_pins,
            "composition": {
                "composition_id": "nationality_criminal_code_offense_composition",
                "label_ja": "日本を含む国籍等別・刑法犯上位6区分の構成",
                "label_en": "Criminal-code offense composition by nationality including Japan",
                "year": 2024,
                "foreign_source_id": "S08",
                "foreign_source_table": "130",
                "all_person_source_id": "S15",
                "all_person_source_table": "3",
                "expected_foreign_country_row_count": 4,
                "expected_foreign_region_total_row_count": 2,
                "aggregated_nationality_label": "アメリカ",
                "expected_aggregated_subcategory_row_count": 2,
                "expected_foreign_entity_count": 5,
                "expected_total_entity_count": 6,
                "category_ids": [item[0] for item in CATEGORIES],
                "category_definitions": [
                    {
                        "offense_id": offense_id,
                        "label_ja": label,
                        "color": color,
                        "display_order": index,
                        "official_severity_role": (
                            "official_high_severity_category"
                            if offense_id == "heinous"
                            else "not_a_project_severity_classification"
                        ),
                    }
                    for index, (offense_id, label, color) in enumerate(CATEGORIES, 1)
                ],
                "metrics": ["cleared_cases", "cleared_persons"],
                "small_number_total_threshold": 20,
                "clustering": {
                    "distance": "jensen_shannon",
                    "log_base": 2,
                    "linkage": "average",
                    "input": "within_entity_composition_share",
                },
                "interpretation_policy": "patterns_without_intrinsic_group_inference",
                "ui_caveat": "構成比は犯罪の多寡や個人riskを示さない。",
            },
        },
    )
    return catalog_path, raw_root, processed_root, contract_path


def test_offense_composition_includes_japanese_residual_and_all_foreign_entities(
    tmp_path,
    nationality_offense_file,
    all_person_offense_file,
):
    catalog, raw_root, processed_root, contract = _build_inputs(
        tmp_path, nationality_offense_file, all_person_offense_file
    )

    report = generate_offense_composition_report(
        catalog_path=catalog,
        raw_root=raw_root,
        processed_root=processed_root,
        contract_path=contract,
        output_root=tmp_path / "output",
        generated_at="2026-09-03T20:30:00+09:00",
    )

    rows = [json.loads(line) for line in report.jsonl_path.read_text().splitlines()]
    summary = json.loads(report.summary_path.read_text())
    latest = json.loads(report.latest_path.read_text())
    japanese = [row for row in rows if row["is_japanese_reference"]]

    assert report.record_count == 36
    assert report.entity_count == 6
    assert len(japanese) == 6
    assert {row["offense_id"] for row in japanese} == {
        item[0] for item in CATEGORIES
    }
    assert {row["criminal_code_cleared_persons_total"] for row in japanese} == {
        120
    }
    assert sum(row["cleared_persons"] for row in japanese) == 120
    assert sum(row["cleared_persons_share"] for row in japanese) == pytest.approx(1)
    assert all(row["derivation_method"] == "residual_subtraction" for row in japanese)

    america = [row for row in rows if row["display_label"] == "アメリカ"]
    assert len(america) == 6
    assert {row["criminal_code_cleared_persons_total"] for row in america} == {40}
    assert all(row["derivation_method"] == "sum_published_subcategories" for row in america)
    assert next(row for row in america if row["offense_id"] == "theft")[
        "cleared_persons"
    ] == 12

    assert summary["status_counts"] == {"calculated": 36, "refused": 0}
    assert summary["category_definitions"][0]["offense_id"] == "heinous"
    assert summary["small_number_total_threshold"] == 20
    assert set(summary["clustering"]) == {"cleared_cases", "cleared_persons"}
    assert set(summary["clustering"]["cleared_persons"]["order"]) == {
        row["entity_id"] for row in rows
    }
    assert len(summary["clustering"]["cleared_persons"]["order"]) == 6
    assert summary["excluded_source_rows"]["region_total_count"] == 2
    assert latest["run_relpath"] == "20260903_203000_offense_composition"
    assert latest["offense_composition_records_sha256"] == sha256_file(
        report.jsonl_path
    )
    assert latest["summary_sha256"] == sha256_file(report.summary_path)


def test_offense_composition_flags_small_totals_without_dropping_rows(
    tmp_path,
    nationality_offense_file,
    all_person_offense_file,
):
    catalog, raw_root, processed_root, contract = _build_inputs(
        tmp_path, nationality_offense_file, all_person_offense_file
    )
    report = generate_offense_composition_report(
        catalog_path=catalog,
        raw_root=raw_root,
        processed_root=processed_root,
        contract_path=contract,
        output_root=tmp_path / "output",
        generated_at="2026-09-03T20:31:00+09:00",
    )
    rows = [json.loads(line) for line in report.jsonl_path.read_text().splitlines()]

    stateless = [row for row in rows if row["published_label"] == "無国籍"]
    assert len(stateless) == 6
    assert all(
        "sparse_entity_total_cleared_persons" in row["small_number_warning_flags"]
        for row in stateless
    )
    assert sum(row["cleared_persons_share"] for row in stateless) == pytest.approx(1)


def test_offense_composition_keeps_metric_that_has_a_nonzero_total(
    tmp_path,
    nationality_offense_file,
    all_person_offense_file,
):
    from openpyxl import load_workbook

    workbook = load_workbook(nationality_offense_file)
    locations = {
        "01 ": (19, (5, 7, 9, 20)),
        "02": (19, (9, 13)),
        "03 ": (20, (7, 16)),
        "04 ": (19, (5,)),
    }
    for sheet_name, (source_row, cases_columns) in locations.items():
        for cases_column in cases_columns:
            workbook[sheet_name].cell(source_row, cases_column, 0)
    workbook.save(nationality_offense_file)
    catalog, raw_root, processed_root, contract = _build_inputs(
        tmp_path, nationality_offense_file, all_person_offense_file
    )

    report = generate_offense_composition_report(
        catalog_path=catalog,
        raw_root=raw_root,
        processed_root=processed_root,
        contract_path=contract,
        output_root=tmp_path / "output",
        generated_at="2026-09-03T20:31:30+09:00",
    )
    rows = [json.loads(line) for line in report.jsonl_path.read_text().splitlines()]
    unknown = [row for row in rows if row["published_label"] == "国籍不明"]

    assert len(unknown) == 6
    assert all(row["calculation_status"] == "calculated" for row in unknown)
    assert all(row["cleared_cases_share_status"] == "refused_zero_total" for row in unknown)
    assert all(row["cleared_cases_share"] is None for row in unknown)
    assert all(row["cleared_persons_share_status"] == "calculated" for row in unknown)
    assert sum(row["cleared_persons_share"] for row in unknown) == pytest.approx(1)


def test_offense_composition_rejects_raw_hash_drift(
    tmp_path,
    nationality_offense_file,
    all_person_offense_file,
):
    catalog, raw_root, processed_root, contract = _build_inputs(
        tmp_path, nationality_offense_file, all_person_offense_file
    )
    contract_data = json.loads(contract.read_text())
    contract_data["artifact_pins"]["S08"] = "f" * 64
    _write_json(contract, contract_data)

    with pytest.raises(IntegrityError, match="artifact pin"):
        generate_offense_composition_report(
            catalog_path=catalog,
            raw_root=raw_root,
            processed_root=processed_root,
            contract_path=contract,
            output_root=tmp_path / "output",
            generated_at="2026-09-03T20:32:00+09:00",
        )


def test_jensen_shannon_distance_handles_zero_shares_and_is_symmetric():
    left = (1.0, 0.0, 0.0)
    right = (0.0, 1.0, 0.0)

    assert jensen_shannon_distance(left, left) == pytest.approx(0)
    assert jensen_shannon_distance(left, right) == pytest.approx(1)
    assert jensen_shannon_distance(left, right) == pytest.approx(
        jensen_shannon_distance(right, left)
    )
