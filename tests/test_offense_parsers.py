import pytest

from nationality_crime_atlas.errors import SchemaError
from nationality_crime_atlas.npa_offenses import (
    TOP_LEVEL_CRIMINAL_CODE_GROUP_IDS,
    parse_npa_all_person_offense_groups,
    parse_npa_nationality_offense_groups,
)


def test_nationality_offense_parser_extracts_official_group_hierarchy(
    nationality_offense_file,
):
    records = parse_npa_nationality_offense_groups(
        nationality_offense_file,
        table_id="130",
        source_id="S08",
    )

    china = [record for record in records if record.nationality == "中国"]
    by_offense = {record.offense_id: record for record in china}

    assert len(china) == 9
    assert tuple(
        offense_id for offense_id in by_offense if offense_id in TOP_LEVEL_CRIMINAL_CODE_GROUP_IDS
    ) == TOP_LEVEL_CRIMINAL_CODE_GROUP_IDS
    assert by_offense["criminal_code"].cleared_persons == 50
    assert sum(
        by_offense[offense_id].cleared_persons
        for offense_id in TOP_LEVEL_CRIMINAL_CODE_GROUP_IDS
    ) == by_offense["criminal_code"].cleared_persons
    assert by_offense["heinous"].cleared_cases == 10
    assert by_offense["heinous"].offense_label == "凶悪犯"
    assert by_offense["heinous"].offense_parent_id == "criminal_code"
    assert by_offense["heinous"].official_severity_role == (
        "official_high_severity_category"
    )
    assert by_offense["theft"].official_severity_role == (
        "not_a_project_severity_classification"
    )
    assert by_offense["theft"].source_sheet == "02"
    assert by_offense["morals"].source_sheet == "03"


def test_nationality_offense_parser_preserves_annual_total_and_subcategories(
    nationality_offense_file,
):
    records = parse_npa_nationality_offense_groups(
        nationality_offense_file,
        table_id="130",
        source_id="S08",
    )

    annual = [record for record in records if record.row_kind == "annual_total"]
    assert len(annual) == 9
    assert {record.year for record in annual} == {2024}
    assert next(
        record for record in annual if record.offense_id == "criminal_code"
    ).cleared_persons == 180

    us_other = next(
        record
        for record in records
        if record.nationality == "アメリカ"
        and record.subcategory == "その他"
        and record.offense_id == "theft"
    )
    assert us_other.row_kind == "subcategory"
    assert us_other.region == "南北アメリカ州の国"
    assert us_other.cleared_persons == 9


def test_nationality_offense_parser_stops_before_verification_rows(
    nationality_offense_file,
):
    from openpyxl import load_workbook

    workbook = load_workbook(nationality_offense_file)
    worksheet = workbook["02"]
    note_row = next(
        cell.row
        for cell in worksheet["B"]
        if isinstance(cell.value, str) and cell.value.startswith("注")
    )
    worksheet.cell(note_row, 2, "３ fixture note without 注 marker")
    worksheet.cell(note_row + 2, 4, "計")
    worksheet.cell(note_row + 2, 9, 0)
    worksheet.cell(note_row + 2, 10, 0)
    worksheet.cell(note_row + 2, 13, 0)
    worksheet.cell(note_row + 2, 14, 0)
    workbook.save(nationality_offense_file)

    records = parse_npa_nationality_offense_groups(
        nationality_offense_file,
        table_id="130",
        source_id="S08_R02",
    )

    assert len(records) == 81
    assert not any(record.subcategory == "計" for record in records)


def test_all_person_parser_extracts_same_six_official_groups(
    all_person_offense_file,
):
    records = parse_npa_all_person_offense_groups(
        all_person_offense_file,
        source_id="S15",
    )

    by_offense = {record.offense_id: record for record in records}
    assert tuple(by_offense) == ("criminal_code",) + TOP_LEVEL_CRIMINAL_CODE_GROUP_IDS
    assert by_offense["heinous"].cleared_persons == 30
    assert by_offense["theft"].cleared_cases == 180
    assert by_offense["criminal_code"].source_sheet == "刑法犯総数"
    assert by_offense["heinous"].source_sheet == "A"
    assert {record.year for record in records} == {2024}


def test_nationality_offense_parser_rejects_missing_required_sheet(
    nationality_offense_file,
):
    from openpyxl import load_workbook

    workbook = load_workbook(nationality_offense_file)
    del workbook["04 "]
    workbook.save(nationality_offense_file)

    with pytest.raises(SchemaError, match="required sheet"):
        parse_npa_nationality_offense_groups(
            nationality_offense_file,
            table_id="130",
            source_id="S08",
        )


def test_all_person_parser_rejects_non_exhaustive_top_level_groups(
    all_person_offense_file,
):
    from openpyxl import load_workbook

    workbook = load_workbook(all_person_offense_file)
    workbook["F"].cell(10, 6, 11)
    workbook.save(all_person_offense_file)

    with pytest.raises(SchemaError, match="do not sum to criminal_code"):
        parse_npa_all_person_offense_groups(
            all_person_offense_file,
            source_id="S15",
        )
